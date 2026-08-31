from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from racktool.cli.main import main
from racktool.core import analyze_workbook
from racktool.profiles import (
    LayoutProfile,
    apply_profile,
    fingerprint_workbook,
    load_profile,
    select_profile,
    validate_profile_data,
)

PROFILE_ROOT = Path(__file__).resolve().parents[2] / "src/racktool/profiles"
DUAL_PROFILE = PROFILE_ROOT / "generic-dual-axis.yaml"
MIXED_PROFILE = PROFILE_ROOT / "generic-mixed-axis.yaml"


def _fill_descending_axis(sheet: Worksheet, column: int, start_row: int, height: int) -> None:
    for offset, u_number in enumerate(range(height, 0, -1)):
        sheet.cell(start_row + offset, column, u_number)


def _make_dual_axis_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "双轴"
    sheet.merge_cells("A1:C1")
    sheet["A1"] = "RACK-A"
    _fill_descending_axis(sheet, 1, 2, 12)
    _fill_descending_axis(sheet, 3, 2, 12)
    sheet.merge_cells("B2:B3")
    sheet["B2"] = "设备 A"
    sheet["B5"] = "空闲"
    workbook.save(path)


def _make_mixed_axis_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "共享轴"
    sheet.merge_cells("A1:B1")
    sheet["A1"] = "RACK-1"
    sheet.merge_cells("C1:D1")
    sheet["C1"] = "RACK-2"
    _fill_descending_axis(sheet, 1, 2, 10)
    _fill_descending_axis(sheet, 3, 2, 10)
    sheet["B2"] = "共享轴左侧设备"
    sheet.merge_cells("D2:D6")
    sheet["D2"] = "边缘机柜设备"
    workbook.save(path)


def test_generic_dual_axis_profile_matches_and_dry_runs(tmp_path: Path) -> None:
    path = tmp_path / "dual-axis.xlsx"
    _make_dual_axis_workbook(path)
    profile = load_profile(DUAL_PROFILE)
    source_bytes = path.read_bytes()

    selection = select_profile(profile, analyze_workbook(path))
    dry_run = apply_profile(path, profile, dry_run=True)
    applied = apply_profile(path, profile, dry_run=False)

    assert selection.status == "matched"
    assert dry_run.status == "applied"
    assert dry_run.dry_run is True
    assert dry_run.analysis is None
    assert applied.status == "applied"
    assert applied.dry_run is False
    assert applied.selected_sheets == ("双轴",)
    assert applied.ignored_device_candidate_ids
    assert path.read_bytes() == source_bytes


def test_wrong_profile_is_rejected_instead_of_guessing(tmp_path: Path) -> None:
    path = tmp_path / "mixed-axis.xlsx"
    _make_mixed_axis_workbook(path)
    dual = load_profile(DUAL_PROFILE)
    mixed = load_profile(MIXED_PROFILE)

    assert select_profile(mixed, analyze_workbook(path)).status == "matched"
    rejected = apply_profile(path, dual, dry_run=True)
    assert rejected.status == "rejected"
    assert rejected.analysis is None


def test_ambiguous_multi_sheet_match_is_not_auto_applied(tmp_path: Path) -> None:
    path = tmp_path / "two-layouts.xlsx"
    workbook = Workbook()
    first = workbook.active
    assert first is not None
    first.title = "布局A"
    first.merge_cells("A1:C1")
    first["A1"] = "RACK-A"
    _fill_descending_axis(first, 1, 2, 12)
    _fill_descending_axis(first, 3, 2, 12)
    first["B2"] = "设备A"
    second = workbook.create_sheet("布局B")
    second.merge_cells("A1:C1")
    second["A1"] = "RACK-B"
    _fill_descending_axis(second, 1, 2, 12)
    _fill_descending_axis(second, 3, 2, 12)
    second["B2"] = "设备B"
    workbook.save(path)

    result = apply_profile(path, load_profile(DUAL_PROFILE), dry_run=True)
    assert result.status == "ambiguous"
    assert result.analysis is None


def test_review_required_profile_needs_explicit_force(tmp_path: Path) -> None:
    path = tmp_path / "numeric-device.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.merge_cells("A1:C1")
    sheet["A1"] = "RACK-NUMERIC"
    _fill_descending_axis(sheet, 1, 2, 12)
    _fill_descending_axis(sheet, 3, 2, 12)
    sheet["B2"] = 10001
    workbook.save(path)

    profile = load_profile(DUAL_PROFILE)
    blocked = apply_profile(path, profile, dry_run=True)
    forced = apply_profile(path, profile, dry_run=True, force=True)

    assert blocked.status == "review_required"
    assert forced.status == "applied"


def test_profile_cli_validate_match_and_dry_run(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    path = tmp_path / "cli-profile.xlsx"
    _make_dual_axis_workbook(path)

    assert main(["profile", "validate", str(DUAL_PROFILE)]) == 0
    validate_out = capsys.readouterr().out
    assert main(["profile", "match", str(path), str(DUAL_PROFILE)]) == 0
    match_out = capsys.readouterr().out
    assert main(["profile", "apply", str(path), str(DUAL_PROFILE)]) == 0
    apply_out = capsys.readouterr().out
    assert json.loads(validate_out)["profile_id"] == "generic-dual-axis"
    assert json.loads(match_out)["status"] == "matched"
    applied = json.loads(apply_out)
    assert applied["status"] == "applied"
    assert applied["dry_run"] is True


def _fixed_coordinate_profile() -> LayoutProfile:
    return validate_profile_data(
        {
            "schema_version": 1,
            "profile_id": "fixed-coordinate-layout",
            "name": "Fixed coordinate layout",
            "match": {"min_confidence": 1.0, "review_confidence": 0.5},
            "rack": {
                "title": {"mode": "fixed_range", "range": "G2:H2"},
                "height": {"mode": "infer_from_u_axis"},
            },
            "u_axis": {
                "direction": "descending",
                "pairing": "paired",
                "left_column": "F",
                "right_column": "I",
                "start_row": 4,
                "end_row": 11,
                "max_missing_rows": 0,
            },
            "device_area": {"mode": "fixed_range", "range": "G4:H11"},
        }
    )


def test_profile_coordinates_parse_layout_default_analyzer_cannot(tmp_path: Path) -> None:
    path = tmp_path / "fixed-coordinate-layout.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "坐标布局"
    sheet.merge_cells("G2:H2")
    sheet["G2"] = "RACK-FIXED"
    _fill_descending_axis(sheet, 6, 4, 8)
    _fill_descending_axis(sheet, 9, 4, 8)
    sheet.merge_cells("G4:H5")
    sheet["G4"] = "设备"
    workbook.save(path)
    source_bytes = path.read_bytes()

    default_analysis = analyze_workbook(path).to_dict()["sheets"][0]
    result = apply_profile(path, _fixed_coordinate_profile(), dry_run=False)

    assert default_analysis["racks"] == []
    assert result.status == "applied"
    assert result.selected_sheets == ("坐标布局",)
    assert result.analysis is not None
    parsed = result.analysis["sheets"][0]
    assert parsed["racks"][0]["rack_name"] == "RACK-FIXED"
    assert parsed["racks"][0]["height_u"] == 8
    assert parsed["placements"][0]["start_u"] == 7
    assert parsed["placements"][0]["end_u"] == 8
    assert path.read_bytes() == source_bytes


def test_review_score_below_threshold_is_rejected(tmp_path: Path) -> None:
    path = tmp_path / "below-review-threshold.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.merge_cells("A1:C1")
    sheet["A1"] = "RACK-NUMERIC"
    _fill_descending_axis(sheet, 1, 2, 12)
    _fill_descending_axis(sheet, 3, 2, 12)
    sheet["B2"] = 10001
    workbook.save(path)
    profile = validate_profile_data(
        {
            "schema_version": 1,
            "profile_id": "strict-review-threshold",
            "name": "Strict review threshold",
            "match": {"min_confidence": 1.0, "review_confidence": 0.95},
            "rack": {
                "title": {"mode": "merged_cell_above_u_axis"},
                "height": {"mode": "infer_from_u_axis"},
            },
            "u_axis": {"direction": "descending", "pairing": "paired"},
            "device_area": {"mode": "between_u_axes"},
        }
    )

    selection = select_profile(profile, analyze_workbook(path))
    forced = apply_profile(path, profile, dry_run=True, force=True)

    assert selection.matches[0].score < profile.match.review_confidence
    assert selection.matches[0].status == "rejected"
    assert selection.status == "unmatched"
    assert forced.status == "rejected"


def test_force_refuses_multiple_review_sheets_without_permission(tmp_path: Path) -> None:
    path = tmp_path / "two-review-layouts.xlsx"
    workbook = Workbook()
    first = workbook.active
    assert first is not None
    first.title = "布局A"
    second = workbook.create_sheet("布局B")
    for index, sheet in enumerate((first, second), start=1):
        sheet.merge_cells("A1:C1")
        sheet["A1"] = f"RACK-{index}"
        _fill_descending_axis(sheet, 1, 2, 12)
        _fill_descending_axis(sheet, 3, 2, 12)
        sheet["B2"] = 10000 + index
    workbook.save(path)
    profile = load_profile(DUAL_PROFILE)

    selection = select_profile(profile, analyze_workbook(path))
    forced = apply_profile(path, profile, dry_run=True, force=True)

    assert selection.status == "ambiguous"
    assert forced.status == "ambiguous"
    assert forced.selected_sheets == ()
    assert forced.analysis is None


def test_ignore_filter_recomputes_overlap_issues(tmp_path: Path) -> None:
    path = tmp_path / "ignored-overlap.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.merge_cells("A1:D1")
    sheet["A1"] = "RACK-WIDE"
    _fill_descending_axis(sheet, 1, 2, 12)
    _fill_descending_axis(sheet, 4, 2, 12)
    sheet["B2"] = "空闲"
    sheet["C2"] = "设备"
    workbook.save(path)
    profile = validate_profile_data(
        {
            "schema_version": 1,
            "profile_id": "ignore-overlap",
            "name": "Ignore overlap",
            "match": {
                "min_confidence": 1.0,
                "review_confidence": 0.5,
                "require_issue_free": True,
            },
            "rack": {
                "title": {"mode": "merged_cell_above_u_axis"},
                "height": {"mode": "infer_from_u_axis"},
            },
            "u_axis": {"direction": "descending", "pairing": "paired"},
            "device_area": {"mode": "between_u_axes"},
            "text": {"ignore_exact": ["空闲"]},
        }
    )

    unfiltered = analyze_workbook(path).to_dict()["sheets"][0]
    result = apply_profile(path, profile, dry_run=False)

    assert [issue["code"] for issue in unfiltered["issues"]] == [
        "overlapping-placement-candidates"
    ]
    assert result.status == "applied"
    assert result.analysis is not None
    parsed = result.analysis["sheets"][0]
    assert [device["display_text"] for device in parsed["devices"]] == ["设备"]
    assert len(parsed["placements"]) == 1
    assert parsed["issues"] == []


def _make_fingerprint_workbook(
    path: Path,
    *,
    left_column: int,
    title_row: int,
    rack_name: str,
    device_text: str,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "相同布局名"
    right_column = left_column + 2
    sheet.merge_cells(
        start_row=title_row,
        start_column=left_column,
        end_row=title_row,
        end_column=right_column,
    )
    sheet.cell(title_row, left_column, rack_name)
    _fill_descending_axis(sheet, left_column, title_row + 1, 8)
    _fill_descending_axis(sheet, right_column, title_row + 1, 8)
    sheet.cell(title_row + 1, left_column + 1, device_text)
    workbook.save(path)


def test_layout_fingerprint_tracks_geometry_not_business_text(tmp_path: Path) -> None:
    first_path = tmp_path / "first.xlsx"
    renamed_path = tmp_path / "renamed.xlsx"
    shifted_path = tmp_path / "shifted.xlsx"
    _make_fingerprint_workbook(
        first_path,
        left_column=1,
        title_row=1,
        rack_name="RACK-A",
        device_text="设备A",
    )
    _make_fingerprint_workbook(
        renamed_path,
        left_column=1,
        title_row=1,
        rack_name="RACK-B",
        device_text="设备B",
    )
    _make_fingerprint_workbook(
        shifted_path,
        left_column=6,
        title_row=10,
        rack_name="RACK-A",
        device_text="设备A",
    )

    first = fingerprint_workbook(first_path)
    repeated = fingerprint_workbook(first_path)
    renamed = fingerprint_workbook(renamed_path)
    shifted = fingerprint_workbook(shifted_path)

    assert first == repeated
    assert first.workbook_sha256 != renamed.workbook_sha256
    assert first.layout_sha256 == renamed.layout_sha256
    assert first.features == renamed.features
    assert first.layout_sha256 != shifted.layout_sha256
    assert first.features != shifted.features

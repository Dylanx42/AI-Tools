from __future__ import annotations

import sqlite3
from copy import copy
from dataclasses import replace
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from racktool.cli.main import main
from racktool.core.project import import_workbook, rescan_workbook, validate_project
from racktool.models.project import RackProject
from racktool.persistence import load_project, save_project


def _fill_descending_axis(
    sheet: Worksheet,
    column: int,
    start_row: int,
    height: int,
) -> None:
    for offset, u_number in enumerate(range(height, 0, -1)):
        sheet.cell(start_row + offset, column, u_number)


def _add_rack(
    sheet: Worksheet,
    start_column: int,
    name: str,
    devices: list[tuple[int, str]],
) -> None:
    sheet.merge_cells(
        start_row=1,
        start_column=start_column,
        end_row=1,
        end_column=start_column + 2,
    )
    sheet.cell(1, start_column, name)
    _fill_descending_axis(sheet, start_column, 2, 12)
    _fill_descending_axis(sheet, start_column + 2, 2, 12)
    for row, text in devices:
        sheet.cell(row, start_column + 1, text)


def _make_layout(
    path: Path,
    left_text: str = "设备 A",
    right_text: str = "设备 B",
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "机柜"
    _add_rack(sheet, 1, "RACK-A", [(2, left_text), (5, right_text)])
    workbook.save(path)


def _make_two_racks(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "机柜"
    _add_rack(sheet, 1, "RACK-A", [(2, "设备 A")])
    _add_rack(sheet, 5, "RACK-B", [(2, "设备 B")])
    workbook.save(path)


def _device_by_text(project: RackProject, text: str):
    return next(device for device in project.devices if device.display_text == text)


def _rack_by_name(project: RackProject, name: str):
    return next(rack for rack in project.racks if rack.rack_name == name)


def _placement_for(project: RackProject, device_id: str):
    return next(item for item in project.placements if item.device_id == device_id)


def _mapping_for(project: RackProject, device_id: str):
    return next(
        item
        for item in project.mappings
        if item.mapping_kind == "device" and item.device_id == device_id
    )


def test_import_device_mapping_includes_rack_identity(tmp_path: Path) -> None:
    path = tmp_path / "layout.xlsx"
    _make_layout(path)

    project = import_workbook(path)

    placements = {item.device_id: item for item in project.placements}
    device_mappings = [
        item for item in project.mappings if item.mapping_kind == "device"
    ]
    assert device_mappings
    assert all(item.rack_id == placements[item.device_id].rack_id for item in device_mappings)
    assert not [item for item in validate_project(project) if item.severity == "error"]


def test_device_range_swap_preserves_both_identities(tmp_path: Path) -> None:
    path = tmp_path / "layout.xlsx"
    _make_layout(path)
    imported = import_workbook(path)
    ids = {item.display_text: item.device_id for item in imported.devices}

    workbook = load_workbook(path)
    sheet = workbook.active
    sheet["B2"], sheet["B5"] = sheet["B5"].value, sheet["B2"].value
    workbook.save(path)

    result = rescan_workbook(path, imported)

    assert result.accepted
    assert {item.display_text: item.device_id for item in result.project.devices} == ids
    assert _mapping_for(result.project, ids["设备 A"]).source_range == "B5"
    assert _mapping_for(result.project, ids["设备 B"]).source_range == "B2"


def test_rename_and_move_uses_unique_style_and_height_evidence(tmp_path: Path) -> None:
    path = tmp_path / "layout.xlsx"
    _make_layout(path)
    workbook = load_workbook(path)
    workbook.active["B2"].font = Font(bold=True, color="FF0000")
    workbook.save(path)
    imported = import_workbook(path)
    original = _device_by_text(imported, "设备 A")

    workbook = load_workbook(path)
    sheet = workbook.active
    old_style = copy(sheet["B2"]._style)
    sheet["B2"] = None
    sheet["B8"] = "设备 A-改名并移动"
    sheet["B8"]._style = old_style
    workbook.save(path)

    result = rescan_workbook(path, imported)

    assert result.accepted
    renamed = next(
        item for item in result.project.devices if item.device_id == original.device_id
    )
    assert renamed.display_text == "设备 A-改名并移动"
    assert _mapping_for(result.project, original.device_id).source_range == "B8"


def test_rack_relocation_preserves_rack_and_device_identities(tmp_path: Path) -> None:
    path = tmp_path / "layout.xlsx"
    _make_two_racks(path)
    imported = import_workbook(path)
    rack_ids = {item.rack_name: item.rack_id for item in imported.racks}
    device_ids = {item.display_text: item.device_id for item in imported.devices}

    workbook = load_workbook(path)
    sheet = workbook.active
    sheet["A1"], sheet["E1"] = sheet["E1"].value, sheet["A1"].value
    sheet["B2"], sheet["F2"] = sheet["F2"].value, sheet["B2"].value
    workbook.save(path)

    result = rescan_workbook(path, imported)

    assert result.accepted
    assert {item.rack_name: item.rack_id for item in result.project.racks} == rack_ids
    assert {item.display_text: item.device_id for item in result.project.devices} == device_ids
    assert _rack_by_name(result.project, "RACK-A").title_range == "E1:G1"
    assert _mapping_for(result.project, device_ids["设备 A"]).source_range == "F2"


def test_rack_title_rename_at_same_range_preserves_rack_id(tmp_path: Path) -> None:
    path = tmp_path / "layout.xlsx"
    _make_layout(path)
    imported = import_workbook(path)
    original_rack_id = imported.racks[0].rack_id

    workbook = load_workbook(path)
    workbook.active["A1"] = "RACK-A-改名"
    workbook.save(path)
    result = rescan_workbook(path, imported)

    assert result.accepted
    assert _rack_by_name(result.project, "RACK-A-改名").rack_id == original_rack_id


def test_rack_removal_retains_references_and_reappearance_reuses_ids(
    tmp_path: Path,
) -> None:
    path = tmp_path / "layout.xlsx"
    _make_two_racks(path)
    imported = import_workbook(path)
    rack_b = _rack_by_name(imported, "RACK-B")
    device_b = _device_by_text(imported, "设备 B")

    workbook = load_workbook(path)
    sheet = workbook.active
    sheet.unmerge_cells("E1:G1")
    for row in range(1, 14):
        for column in range(5, 8):
            sheet.cell(row, column).value = None
    workbook.save(path)

    missing = rescan_workbook(path, imported)

    assert missing.accepted
    assert rack_b.rack_id in missing.missing_rack_ids
    assert device_b.device_id in missing.missing_device_ids
    assert next(
        item for item in missing.project.racks if item.rack_id == rack_b.rack_id
    ).status == "missing"
    missing_placement = _placement_for(missing.project, device_b.device_id)
    assert missing_placement.status == "missing"
    assert missing_placement.rack_id == rack_b.rack_id
    assert not [item for item in validate_project(missing.project) if item.severity == "error"]

    workbook = load_workbook(path)
    _add_rack(workbook.active, 5, "RACK-B", [(2, "设备 B")])
    workbook.save(path)
    reappeared = rescan_workbook(path, missing.project)

    assert reappeared.accepted
    assert _rack_by_name(reappeared.project, "RACK-B").rack_id == rack_b.rack_id
    assert _device_by_text(reappeared.project, "设备 B").device_id == device_b.device_id
    assert _placement_for(reappeared.project, device_b.device_id).status == "active"


def test_ambiguous_rescan_fails_closed_without_mutating_project(tmp_path: Path) -> None:
    path = tmp_path / "layout.xlsx"
    _make_layout(path, "同名设备", "同名设备")
    imported = import_workbook(path)

    workbook = load_workbook(path)
    sheet = workbook.active
    sheet["B2"] = None
    sheet["B5"] = None
    sheet["B8"] = "同名设备"
    sheet["B10"] = "同名设备"
    workbook.save(path)

    result = rescan_workbook(path, imported)

    assert result.accepted is False
    assert result.project == imported
    assert result.created_device_ids == ()
    assert result.missing_device_ids == ()
    assert any(item.code == "ambiguous-device-identity" for item in result.conflicts)


def test_cli_ambiguous_rescan_returns_failure_without_persisting(
    tmp_path: Path,
    capsys: pytest.CaptureFixture[str],
) -> None:
    path = tmp_path / "layout.xlsx"
    database = tmp_path / "project.sqlite"
    _make_layout(path, "同名设备", "同名设备")
    assert main(["project", "import", str(path), str(database)]) == 0
    capsys.readouterr()
    imported = load_project(database)

    workbook = load_workbook(path)
    sheet = workbook.active
    sheet["B2"] = None
    sheet["B5"] = None
    sheet["B8"] = "同名设备"
    sheet["B10"] = "同名设备"
    workbook.save(path)

    assert main(["project", "rescan", str(path), str(database)]) == 2
    output = capsys.readouterr().out

    assert '"accepted": false' in output
    assert load_project(database).to_dict() == imported.to_dict()


def test_rescan_rejects_different_changed_workbook(tmp_path: Path) -> None:
    path = tmp_path / "layout.xlsx"
    other = tmp_path / "other.xlsx"
    _make_layout(path)
    _make_layout(other, "其他 A", "其他 B")
    imported = import_workbook(path)

    result = rescan_workbook(other, imported)

    assert result.accepted is False
    assert result.project == imported
    assert any(item.code == "wrong-source-workbook" for item in result.conflicts)


def test_sqlite_round_trip_preserves_identities_and_normalizes_source(
    tmp_path: Path,
) -> None:
    path = tmp_path / "layout.xlsx"
    database = tmp_path / "project.sqlite"
    _make_layout(path)
    imported = import_workbook(path)

    save_project(database, replace(imported, source_workbook=str(path)))
    loaded = load_project(database)

    assert loaded.project_id == imported.project_id
    assert loaded.source_workbook == str(path.resolve())
    assert {device.device_id for device in loaded.devices} == {
        device.device_id for device in imported.devices
    }


def test_sqlite_reuse_atomically_replaces_prior_project(tmp_path: Path) -> None:
    first_path = tmp_path / "first.xlsx"
    second_path = tmp_path / "second.xlsx"
    database = tmp_path / "project.sqlite"
    _make_layout(first_path)
    _make_layout(second_path, "新设备 A", "新设备 B")
    first = import_workbook(first_path)
    second = import_workbook(second_path)

    save_project(database, first)
    save_project(database, second)
    loaded = load_project(database)

    assert loaded.project_id == second.project_id
    assert {item.display_text for item in loaded.devices} == {"新设备 A", "新设备 B"}
    connection = sqlite3.connect(database)
    try:
        assert connection.execute("SELECT COUNT(*) FROM projects").fetchone() == (1,)
        assert connection.execute(
            "SELECT COUNT(*) FROM devices WHERE project_id = ?",
            (first.project_id,),
        ).fetchone() == (0,)
    finally:
        connection.close()


def test_sqlite_save_failure_leaves_previous_database_readable(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    first_path = tmp_path / "first.xlsx"
    second_path = tmp_path / "second.xlsx"
    database = tmp_path / "project.sqlite"
    _make_layout(first_path)
    _make_layout(second_path, "新设备 A", "新设备 B")
    first = import_workbook(first_path)
    second = import_workbook(second_path)
    save_project(database, first)

    def fail_replace(source: Path, target: Path) -> None:
        raise OSError(f"forced replace failure: {source} -> {target}")

    monkeypatch.setattr("racktool.persistence.sqlite.os.replace", fail_replace)
    with pytest.raises(OSError, match="forced replace failure"):
        save_project(database, second)

    assert load_project(database).project_id == first.project_id
    assert not list(tmp_path.glob(".project.sqlite.tmp-*"))

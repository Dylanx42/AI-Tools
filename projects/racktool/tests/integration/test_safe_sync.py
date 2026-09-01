from __future__ import annotations

from dataclasses import replace
from pathlib import Path
from zipfile import ZipFile

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from racktool.cli.main import main
from racktool.core.backup import create_backup, restore_backup
from racktool.core.project import import_workbook
from racktool.core.service import commit_write_plan, import_project
from racktool.core.sync import apply_writeback, plan_device_move
from racktool.models.project import IdentityConflict
from racktool.persistence import load_project


def _fill_descending_axis(
    sheet: Worksheet,
    column: int,
    start_row: int,
    height: int,
) -> None:
    for offset, u_number in enumerate(range(height, 0, -1)):
        sheet.cell(start_row + offset, column, u_number)


def _make_layout(path: Path, *, blank_target_merge: bool = False) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "机柜"
    sheet.merge_cells("A1:C1")
    sheet["A1"] = "RACK-A"
    _fill_descending_axis(sheet, 1, 2, 12)
    _fill_descending_axis(sheet, 3, 2, 12)
    sheet.merge_cells("B2:B3")
    sheet["B2"] = "设备 A"
    sheet["B2"].font = Font(bold=True)
    sheet["B2"].fill = PatternFill("solid", fgColor="00FF00")
    sheet["B5"] = "设备 B"
    if blank_target_merge:
        sheet.merge_cells("B9:B10")
    sheet["Z1"] = "保留备注"
    workbook.create_sheet("无关表").append(["不要删除"])
    workbook.save(path)


def _make_multicolumn_layout(path: Path, *, merge_device: bool = True) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "机柜"
    sheet.merge_cells("A1:D1")
    sheet["A1"] = "RACK-WIDE"
    _fill_descending_axis(sheet, 1, 2, 12)
    _fill_descending_axis(sheet, 4, 2, 12)
    if merge_device:
        sheet.merge_cells("B2:C2")
    sheet["B2"] = "双列设备"
    workbook.save(path)


def _device(project, text: str):
    return next(item for item in project.devices if item.display_text == text)


def test_normal_move_updates_1u_and_preserves_unrelated_data(tmp_path: Path) -> None:
    path = tmp_path / "layout.xlsx"
    _make_layout(path)
    original = path.read_bytes()
    project = import_workbook(path)
    device = _device(project, "设备 B")
    rack_id = project.racks[0].rack_id

    plan = plan_device_move(
        project,
        device.device_id,
        rack_id,
        4,
        4,
        workbook_path=path,
    )
    result = apply_writeback(path, project, plan)

    assert result.status == "applied"
    assert result.backup_path is not None
    assert Path(result.backup_path).read_bytes() == original
    assert result.project is not None
    assert result.project.source_workbook == str(path.resolve())
    assert Path(result.project.source_workbook).is_file()
    workbook = load_workbook(path)
    try:
        assert workbook.active["B10"].value == "设备 B"
        assert workbook.active["Z1"].value == "保留备注"
        assert [sheet.title for sheet in workbook.worksheets] == ["机柜", "无关表"]
    finally:
        workbook.close()


def test_multi_u_move_preserves_height_style_and_merge_shape(tmp_path: Path) -> None:
    path = tmp_path / "layout.xlsx"
    _make_layout(path)
    project = import_workbook(path)
    device = _device(project, "设备 A")
    rack_id = project.racks[0].rack_id

    plan = plan_device_move(
        project,
        device.device_id,
        rack_id,
        5,
        6,
        workbook_path=path,
    )
    result = apply_writeback(path, project, plan)

    assert result.status == "applied"
    workbook = load_workbook(path)
    try:
        merged = {str(item).replace("$", "") for item in workbook.active.merged_cells.ranges}
        assert "B8:B9" in merged
        assert workbook.active["B8"].value == "设备 A"
        assert workbook.active["B8"].font.bold is True
        assert workbook.active["B2"].value in (None, "")
    finally:
        workbook.close()


@pytest.mark.parametrize("coordinate", ["B8", "B9"])
def test_target_comment_blocks_multi_u_move_before_write(
    tmp_path: Path,
    coordinate: str,
) -> None:
    path = tmp_path / "layout.xlsx"
    _make_layout(path)
    workbook = load_workbook(path)
    workbook.active[coordinate].comment = Comment("unknown note", "audit")
    workbook.save(path)
    workbook.close()
    original = path.read_bytes()
    project = import_workbook(path)
    device = _device(project, "设备 A")
    rack_id = project.racks[0].rack_id

    plan = plan_device_move(
        project,
        device.device_id,
        rack_id,
        5,
        6,
        workbook_path=path,
    )
    result = apply_writeback(path, project, plan)

    assert any(item.code == "unsupported-cell-comment" for item in plan.conflicts)
    assert not any(item.code == "unsupported-threaded-comments" for item in plan.conflicts)
    assert result.status == "rejected"
    assert result.backup_path is None
    assert result.project == project
    assert path.read_bytes() == original
    assert list(tmp_path.glob("layout.xlsx.bak-*")) == []


@pytest.mark.parametrize(
    "package_part",
    [
        "xl/threadedComments/threadedComment1.xml",
        "xl/persons/person.xml",
    ],
)
def test_threaded_comment_package_parts_block_before_openpyxl_load(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    package_part: str,
) -> None:
    path = tmp_path / "layout.xlsx"
    _make_layout(path)
    with ZipFile(path, "a") as package:
        package.writestr(package_part, b"<unsupported-threaded-comment-marker />")
    original = path.read_bytes()
    project = import_workbook(path)
    device = _device(project, "设备 A")
    rack_id = project.racks[0].rack_id

    def unexpected_openpyxl_load(*args, **kwargs):
        raise AssertionError("Safe Sync must inspect package markers before openpyxl load")

    monkeypatch.setattr("racktool.core.sync.load_workbook", unexpected_openpyxl_load)
    plan = plan_device_move(
        project,
        device.device_id,
        rack_id,
        5,
        6,
        workbook_path=path,
    )
    result = apply_writeback(path, project, plan)

    conflicts = [
        item for item in plan.conflicts if item.code == "unsupported-threaded-comments"
    ]
    assert len(conflicts) == 1
    assert f"part={package_part}" in conflicts[0].evidence
    assert result.status == "rejected"
    assert result.backup_path is None
    assert result.project == project
    assert path.read_bytes() == original
    assert list(tmp_path.glob("layout.xlsx.bak-*")) == []


@pytest.mark.parametrize("coordinate", ["B8", "B9"])
def test_target_hyperlink_blocks_multi_u_move_before_write(
    tmp_path: Path,
    coordinate: str,
) -> None:
    path = tmp_path / "layout.xlsx"
    _make_layout(path)
    workbook = load_workbook(path)
    workbook.active[coordinate] = " "
    workbook.active[coordinate].hyperlink = "https://example.invalid/unknown"
    workbook.save(path)
    workbook.close()
    original = path.read_bytes()
    project = import_workbook(path)
    device = _device(project, "设备 A")
    rack_id = project.racks[0].rack_id

    plan = plan_device_move(
        project,
        device.device_id,
        rack_id,
        5,
        6,
        workbook_path=path,
    )
    result = apply_writeback(path, project, plan)

    assert any(item.code == "unsupported-cell-hyperlink" for item in plan.conflicts)
    assert result.status == "rejected"
    assert result.backup_path is None
    assert path.read_bytes() == original


@pytest.mark.parametrize("metadata_kind", ["comment", "hyperlink"])
def test_source_annotation_blocks_move_before_write(
    tmp_path: Path,
    metadata_kind: str,
) -> None:
    path = tmp_path / "layout.xlsx"
    _make_layout(path)
    workbook = load_workbook(path)
    if metadata_kind == "comment":
        workbook.active["B2"].comment = Comment("device note", "audit")
    else:
        workbook.active["B2"].hyperlink = "https://example.invalid/device"
    workbook.save(path)
    workbook.close()
    original = path.read_bytes()
    project = import_workbook(path)
    device = _device(project, "设备 A")
    rack_id = project.racks[0].rack_id

    plan = plan_device_move(
        project,
        device.device_id,
        rack_id,
        5,
        6,
        workbook_path=path,
    )
    result = apply_writeback(path, project, plan)

    assert any(
        item.code == f"unsupported-cell-{metadata_kind}" for item in plan.conflicts
    )
    assert result.status == "rejected"
    assert result.backup_path is None
    assert path.read_bytes() == original


@pytest.mark.parametrize("metadata_kind", ["comment", "hyperlink"])
def test_source_non_anchor_annotation_blocks_move_before_write(
    tmp_path: Path,
    metadata_kind: str,
) -> None:
    path = tmp_path / "wide.xlsx"
    _make_multicolumn_layout(path, merge_device=False)
    workbook = load_workbook(path)
    if metadata_kind == "comment":
        workbook.active["C2"].comment = Comment("lane note", "audit")
    else:
        workbook.active["C2"] = " "
        workbook.active["C2"].hyperlink = "https://example.invalid/lane"
    workbook.save(path)
    workbook.close()
    original = path.read_bytes()
    project = import_workbook(path)
    device = _device(project, "双列设备")
    mapping = next(
        item
        for item in project.mappings
        if item.mapping_kind == "device" and item.device_id == device.device_id
    )
    project = replace(
        project,
        mappings=[
            replace(item, source_range="B2:C2")
            if item.mapping_id == mapping.mapping_id
            else item
            for item in project.mappings
        ],
    )
    rack_id = project.racks[0].rack_id

    plan = plan_device_move(
        project,
        device.device_id,
        rack_id,
        4,
        4,
        workbook_path=path,
    )
    result = apply_writeback(path, project, plan)

    assert any(
        item.code == f"unsupported-cell-{metadata_kind}" for item in plan.conflicts
    )
    assert result.status == "rejected"
    assert result.backup_path is None
    assert path.read_bytes() == original


@pytest.mark.parametrize("validation_range", ["B2:B3", "B8:B9"])
def test_data_validation_on_action_range_blocks_move(
    tmp_path: Path,
    validation_range: str,
) -> None:
    path = tmp_path / "layout.xlsx"
    _make_layout(path)
    workbook = load_workbook(path)
    validation = DataValidation(type="list", formula1='"allowed,values"')
    workbook.active.add_data_validation(validation)
    validation.add(validation_range)
    workbook.save(path)
    workbook.close()
    original = path.read_bytes()
    project = import_workbook(path)
    device = _device(project, "设备 A")
    rack_id = project.racks[0].rack_id

    plan = plan_device_move(
        project,
        device.device_id,
        rack_id,
        5,
        6,
        workbook_path=path,
    )
    result = apply_writeback(path, project, plan)

    assert any(item.code == "unsupported-data-validation" for item in plan.conflicts)
    assert result.status == "rejected"
    assert result.backup_path is None
    assert path.read_bytes() == original


@pytest.mark.parametrize("defined_range", ["$B$2:$B$3", "$B$8:$B$9"])
def test_defined_name_on_action_range_blocks_move(
    tmp_path: Path,
    defined_range: str,
) -> None:
    path = tmp_path / "layout.xlsx"
    _make_layout(path)
    workbook = load_workbook(path)
    workbook.defined_names.add(
        DefinedName(
            "action_region",
            attr_text=f"'{workbook.active.title}'!{defined_range}",
        )
    )
    workbook.save(path)
    workbook.close()
    original = path.read_bytes()
    project = import_workbook(path)
    device = _device(project, "设备 A")
    rack_id = project.racks[0].rack_id

    plan = plan_device_move(
        project,
        device.device_id,
        rack_id,
        5,
        6,
        workbook_path=path,
    )
    result = apply_writeback(path, project, plan)

    assert any(item.code == "unsupported-defined-name" for item in plan.conflicts)
    assert result.status == "rejected"
    assert result.backup_path is None
    assert path.read_bytes() == original


def test_worksheet_scoped_defined_name_on_action_range_blocks_move(
    tmp_path: Path,
) -> None:
    path = tmp_path / "layout.xlsx"
    _make_layout(path)
    workbook = load_workbook(path)
    workbook.active.defined_names.add(
        DefinedName("local_action_region", attr_text="$B$8:$B$9")
    )
    workbook.save(path)
    workbook.close()
    original = path.read_bytes()
    project = import_workbook(path)
    device = _device(project, "设备 A")
    rack_id = project.racks[0].rack_id

    plan = plan_device_move(
        project,
        device.device_id,
        rack_id,
        5,
        6,
        workbook_path=path,
    )
    result = apply_writeback(path, project, plan)

    assert any(item.code == "unsupported-defined-name" for item in plan.conflicts)
    assert result.status == "rejected"
    assert result.backup_path is None
    assert path.read_bytes() == original


def test_metadata_outside_action_ranges_is_preserved(tmp_path: Path) -> None:
    path = tmp_path / "layout.xlsx"
    _make_layout(path)
    workbook = load_workbook(path)
    workbook.active["Z1"].comment = Comment("keep this note", "audit")
    workbook.active["Z1"].hyperlink = "https://example.invalid/keep"
    validation = DataValidation(
        type="list",
        formula1='"keep,values"',
        allow_blank=True,
    )
    workbook.active.add_data_validation(validation)
    validation.add("Z2:Z3")
    workbook.defined_names.add(
        DefinedName(
            "unrelated_region",
            attr_text=f"'{workbook.active.title}'!$Z$1:$Z$3",
        )
    )
    workbook.save(path)
    workbook.close()
    project = import_workbook(path)
    device = _device(project, "设备 A")
    rack_id = project.racks[0].rack_id

    plan = plan_device_move(
        project,
        device.device_id,
        rack_id,
        5,
        6,
        workbook_path=path,
    )
    result = apply_writeback(path, project, plan)

    assert not plan.conflicts
    assert result.status == "applied"
    reloaded = load_workbook(path)
    try:
        assert reloaded.active["Z1"].comment is not None
        assert reloaded.active["Z1"].comment.text == "keep this note"
        assert reloaded.active["Z1"].comment.author == "audit"
        assert reloaded.active["Z1"].hyperlink is not None
        assert reloaded.active["Z1"].hyperlink.target == "https://example.invalid/keep"
        validations = list(reloaded.active.data_validations.dataValidation)
        assert len(validations) == 1
        assert str(validations[0].sqref) == "Z2:Z3"
        assert validations[0].formula1 == '"keep,values"'
        assert validations[0].allow_blank is True
        defined_name = reloaded.defined_names["unrelated_region"]
        assert defined_name.attr_text == "'机柜'!$Z$1:$Z$3"
        assert list(defined_name.destinations) == [("机柜", "$Z$1:$Z$3")]
    finally:
        reloaded.close()


def test_height_change_is_rejected_without_changing_source(tmp_path: Path) -> None:
    path = tmp_path / "layout.xlsx"
    _make_layout(path)
    original = path.read_bytes()
    project = import_workbook(path)
    device = _device(project, "设备 A")
    rack_id = project.racks[0].rack_id

    plan = plan_device_move(
        project,
        device.device_id,
        rack_id,
        6,
        8,
        workbook_path=path,
    )
    result = apply_writeback(path, project, plan)

    assert any(item.code == "device-height-change" for item in plan.conflicts)
    assert result.status == "rejected"
    assert path.read_bytes() == original


def test_occupied_and_out_of_bounds_targets_are_rejected(tmp_path: Path) -> None:
    path = tmp_path / "layout.xlsx"
    _make_layout(path)
    original = path.read_bytes()
    project = import_workbook(path)
    device = _device(project, "设备 B")
    rack_id = project.racks[0].rack_id

    occupied = plan_device_move(
        project,
        device.device_id,
        rack_id,
        12,
        12,
        workbook_path=path,
    )
    outside = plan_device_move(
        project,
        device.device_id,
        rack_id,
        13,
        13,
        workbook_path=path,
    )

    assert any(item.code == "target-u-occupied" for item in occupied.conflicts)
    assert any(item.code == "u-out-of-bounds" for item in outside.conflicts)
    assert apply_writeback(path, project, occupied).status == "rejected"
    assert apply_writeback(path, project, outside).status == "rejected"
    assert path.read_bytes() == original


def test_existing_project_errors_and_mapping_ambiguity_block_planning(
    tmp_path: Path,
) -> None:
    path = tmp_path / "layout.xlsx"
    _make_layout(path)
    project = import_workbook(path)
    device = _device(project, "设备 B")
    rack_id = project.racks[0].rack_id
    conflicted = replace(
        project,
        conflicts=[
            IdentityConflict(
                code="preexisting-error",
                severity="error",
                message="project is already unsafe",
            )
        ],
    )
    mapping = next(
        item
        for item in project.mappings
        if item.mapping_kind == "device" and item.device_id == device.device_id
    )
    ambiguous = replace(
        project,
        mappings=[*project.mappings, replace(mapping, mapping_id="MAP-DUPLICATE")],
    )

    error_plan = plan_device_move(
        conflicted,
        device.device_id,
        rack_id,
        4,
        4,
        workbook_path=path,
    )
    ambiguous_plan = plan_device_move(
        ambiguous,
        device.device_id,
        rack_id,
        4,
        4,
        workbook_path=path,
    )

    assert any(item.code == "preexisting-error" for item in error_plan.conflicts)
    assert any(
        item.code in {"ambiguous-device-mapping", "ambiguous-project-entity"}
        for item in ambiguous_plan.conflicts
    )


def test_wrong_workbook_and_stale_sha_are_rejected(tmp_path: Path) -> None:
    path = tmp_path / "layout.xlsx"
    other = tmp_path / "other.xlsx"
    _make_layout(path)
    _make_layout(other)
    project = import_workbook(path)
    device = _device(project, "设备 B")
    rack_id = project.racks[0].rack_id
    plan = plan_device_move(
        project,
        device.device_id,
        rack_id,
        4,
        4,
        workbook_path=path,
    )
    other_original = other.read_bytes()

    wrong_result = apply_writeback(other, project, plan)

    assert wrong_result.status == "rejected"
    assert other.read_bytes() == other_original
    workbook = load_workbook(path)
    workbook.active["Z1"] = "导入后外部修改"
    workbook.save(path)
    externally_modified = path.read_bytes()

    stale_result = apply_writeback(path, project, plan)

    assert stale_result.status == "rejected"
    assert any(
        item.code in {"stale-project-fingerprint", "stale-write-plan"}
        for item in stale_result.plan.conflicts
    )
    assert path.read_bytes() == externally_modified


def test_blank_merged_target_is_refused_and_preserved(tmp_path: Path) -> None:
    path = tmp_path / "layout.xlsx"
    _make_layout(path, blank_target_merge=True)
    original = path.read_bytes()
    project = import_workbook(path)
    device = _device(project, "设备 B")
    rack_id = project.racks[0].rack_id

    plan = plan_device_move(
        project,
        device.device_id,
        rack_id,
        4,
        4,
        workbook_path=path,
    )

    assert any(item.code == "target-merge-conflict" for item in plan.conflicts)
    assert apply_writeback(path, project, plan).status == "rejected"
    assert path.read_bytes() == original


def test_multi_column_device_width_is_preserved(tmp_path: Path) -> None:
    path = tmp_path / "wide.xlsx"
    _make_multicolumn_layout(path)
    project = import_workbook(path)
    device = _device(project, "双列设备")
    rack_id = project.racks[0].rack_id

    plan = plan_device_move(
        project,
        device.device_id,
        rack_id,
        4,
        4,
        workbook_path=path,
    )
    result = apply_writeback(path, project, plan)

    assert not plan.conflicts
    assert plan.actions[0].old_range == "B2:C2"
    assert plan.actions[0].new_range == "B10:C10"
    assert result.status == "applied"
    workbook = load_workbook(path)
    try:
        assert workbook.active["B10"].value == "双列设备"
        assert "B10:C10" in {
            str(item).replace("$", "") for item in workbook.active.merged_cells.ranges
        }
    finally:
        workbook.close()


def test_failed_temp_validation_leaves_original_file(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = tmp_path / "layout.xlsx"
    _make_layout(path)
    original = path.read_bytes()
    project = import_workbook(path)
    device = _device(project, "设备 B")
    rack_id = project.racks[0].rack_id
    plan = plan_device_move(
        project,
        device.device_id,
        rack_id,
        4,
        4,
        workbook_path=path,
    )

    def boom(*args, **kwargs):
        raise ValueError("forced validation failure")

    monkeypatch.setattr("racktool.core.sync._validate_written_workbook", boom)
    result = apply_writeback(path, project, plan)

    assert result.status == "failed"
    assert path.read_bytes() == original


def test_backups_are_unique_and_damaged_restore_is_fail_safe(tmp_path: Path) -> None:
    path = tmp_path / "layout.xlsx"
    _make_layout(path)
    original = path.read_bytes()

    first = create_backup(path)
    second = create_backup(path)

    assert first != second
    assert first.is_file() and second.is_file()
    first.write_bytes(b"not an xlsx")
    with pytest.raises(Exception):  # noqa: B017
        restore_backup(first, path)
    assert path.read_bytes() == original


def test_database_save_failure_rolls_back_workbook_and_project(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = tmp_path / "layout.xlsx"
    database = tmp_path / "project.sqlite"
    _make_layout(path)
    original_workbook = path.read_bytes()
    original_project = import_project(path, database)
    device = _device(original_project, "设备 B")
    rack_id = original_project.racks[0].rack_id
    plan = plan_device_move(
        original_project,
        device.device_id,
        rack_id,
        4,
        4,
        workbook_path=path,
    )

    def fail_save(*args, **kwargs):
        raise OSError("forced database save failure")

    monkeypatch.setattr("racktool.core.service.save_project", fail_save)
    result = commit_write_plan(path, database, plan)

    assert result.status == "failed"
    assert path.read_bytes() == original_workbook
    assert load_project(database).to_dict() == original_project.to_dict()


def test_cli_sync_commit_updates_workbook_and_sqlite_together(
    tmp_path: Path,
    capsys: pytest.CaptureFixture[str],
) -> None:
    path = tmp_path / "layout.xlsx"
    database = tmp_path / "project.sqlite"
    _make_layout(path)
    assert main(["project", "import", str(path), str(database)]) == 0
    capsys.readouterr()
    project = load_project(database)
    device = _device(project, "设备 B")
    rack_id = project.racks[0].rack_id

    assert (
        main(
            [
                "sync",
                "move",
                str(path),
                str(database),
                device.device_id,
                rack_id,
                "4",
                "4",
                "--commit",
            ]
        )
        == 0
    )
    output = capsys.readouterr().out

    assert '"status": "applied"' in output
    persisted = load_project(database)
    mapping = next(
        item
        for item in persisted.mappings
        if item.mapping_kind == "device" and item.device_id == device.device_id
    )
    assert mapping.source_range == "B10"
    workbook = load_workbook(path)
    try:
        assert workbook.active["B10"].value == "设备 B"
    finally:
        workbook.close()

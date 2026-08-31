from pathlib import Path
from zipfile import ZipFile

import pytest
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from racktool.core import scan_workbook


def _make_synthetic_workbook(path: Path) -> None:
    workbook = Workbook()
    single = workbook.active
    assert single is not None
    single.title = "单轴 12U"
    single.column_dimensions["A"].width = 8
    single.column_dimensions["B"].width = 24
    single.row_dimensions[2].height = 22
    single["A1"] = "U"
    single["B1"] = "设备"
    for offset, u_number in enumerate(range(12, 0, -1), start=2):
        single.cell(offset, 1, u_number)
    single.merge_cells("B2:B2")  # 1U mechanics
    single["B2"] = "交换机\n核心"
    single.merge_cells("B3:B4")  # 2U mechanics
    single["B3"] = "服务器 A"
    single.merge_cells("B5:B8")  # multi-U mechanics
    single["B5"] = "存储阵列\n四行占位"
    single["B3"].fill = PatternFill("solid", fgColor="00FF00")

    dual = workbook.create_sheet("双轴 10U")
    dual.append(["左 U", "设备区域", "右 U"])
    for row, u_number in enumerate(range(10, 0, -1), start=2):
        dual.cell(row, 1, u_number)
        dual.cell(row, 3, u_number)
    dual.merge_cells("B2:B4")
    dual["B2"] = "防火墙\n三行文本"
    workbook.save(path)


def test_scanner_covers_synthetic_structure_and_is_deterministic(tmp_path: Path) -> None:
    path = tmp_path / "synthetic.xlsx"
    _make_synthetic_workbook(path)
    source_bytes = path.read_bytes()

    first = scan_workbook(path).to_dict()
    second = scan_workbook(path).to_dict()

    assert first == second
    assert path.read_bytes() == source_bytes
    assert [sheet["name"] for sheet in first["sheets"]] == ["单轴 12U", "双轴 10U"]
    assert first["sheets"][0]["used_range"] == "A1:B13"
    assert first["sheets"][0]["merged_ranges"] == ["B2", "B3:B4", "B5:B8"]
    assert first["sheets"][0]["row_heights"] == {"2": 22.0}
    assert first["sheets"][0]["column_widths"] == {"A": 8.0, "B": 24.0}
    assert first["sheets"][1]["merged_ranges"] == ["B2:B4"]
    assert any(cell["value"] == "交换机\n核心" for cell in first["sheets"][0]["cells"])
    cells = {cell["coordinate"]: cell for cell in first["sheets"][0]["cells"]}
    assert len(cells["B2"]["style_signature"]) == 16
    assert cells["B2"]["style_signature"] != cells["B3"]["style_signature"]


def test_scanner_rejects_non_xlsx(tmp_path: Path) -> None:
    path = tmp_path / "not-a-workbook.txt"
    path.write_text("not xlsx", encoding="utf-8")
    with pytest.raises(ValueError, match="only .xlsx"):
        scan_workbook(path)


def test_scanner_rejects_malformed_xlsx(tmp_path: Path) -> None:
    path = tmp_path / "malformed.xlsx"
    path.write_bytes(b"not an OOXML archive")

    with pytest.raises(ValueError, match="Invalid XLSX workbook"):
        scan_workbook(path)


def test_scanner_preserves_styled_blank_cells_without_expanding_content_range(
    tmp_path: Path,
) -> None:
    path = tmp_path / "styled-blank.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet["A1"] = "content"
    sheet["C3"].fill = PatternFill("solid", fgColor="00FF00")
    workbook.save(path)
    source_bytes = path.read_bytes()

    result = scan_workbook(path).to_dict()["sheets"][0]
    cells = {cell["coordinate"]: cell for cell in result["cells"]}

    assert result["used_range"] == "A1:A1"
    assert cells["C3"]["value"] is None
    assert cells["C3"]["style_signature"] != cells["A1"]["style_signature"]
    assert path.read_bytes() == source_bytes


def test_scanner_normalizes_missing_ooxml_manifest(tmp_path: Path) -> None:
    path = tmp_path / "missing-manifest.xlsx"
    with ZipFile(path, "w") as archive:
        archive.writestr("placeholder.txt", "not an OOXML workbook")
    source_bytes = path.read_bytes()

    with pytest.raises(ValueError, match="Invalid XLSX workbook"):
        scan_workbook(path)

    assert path.read_bytes() == source_bytes


def test_scanner_normalizes_malformed_ooxml_xml(tmp_path: Path) -> None:
    valid_path = tmp_path / "valid.xlsx"
    workbook = Workbook()
    workbook.save(valid_path)
    with ZipFile(valid_path) as archive:
        entries = {
            info.filename: archive.read(info.filename)
            for info in archive.infolist()
        }

    path = tmp_path / "malformed-xml.xlsx"
    entries["xl/workbook.xml"] = b"<workbook>"
    with ZipFile(path, "w") as archive:
        for name, payload in entries.items():
            archive.writestr(name, payload)
    source_bytes = path.read_bytes()

    with pytest.raises(ValueError, match="Invalid XLSX workbook"):
        scan_workbook(path)

    assert path.read_bytes() == source_bytes

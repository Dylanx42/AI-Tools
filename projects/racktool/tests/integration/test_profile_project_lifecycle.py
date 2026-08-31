from __future__ import annotations

import json
import sqlite3
from dataclasses import replace
from pathlib import Path

import pytest
import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from racktool.cli.main import main
from racktool.core import analyze_workbook, import_workbook
from racktool.core.service import commit_write_plan, import_project, rescan_project
from racktool.core.sync import plan_device_move
from racktool.gui.session import GuiSession
from racktool.persistence import load_project
from racktool.profiles import (
    PROFILE_METADATA_KEY,
    ProfileValidationError,
    canonical_profile_payload,
    load_stored_profile,
    make_profile_record,
    validate_profile_data,
)


def _fill_descending_axis(
    sheet: Worksheet,
    column: int,
    start_row: int,
    height: int,
) -> None:
    for offset, u_number in enumerate(range(height, 0, -1)):
        sheet.cell(start_row + offset, column, u_number)


def _make_fixed_coordinate_layout(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "坐标布局"
    sheet.merge_cells("G2:H2")
    sheet["G2"] = "RACK-FIXED"
    _fill_descending_axis(sheet, 6, 4, 8)
    _fill_descending_axis(sheet, 9, 4, 8)
    sheet.merge_cells("G4:H5")
    sheet["G4"] = "设备 A"
    sheet["Z1"] = "无关内容"
    workbook.save(path)


def _fixed_coordinate_profile():
    return validate_profile_data(
        {
            "schema_version": 1,
            "profile_id": "fixed-coordinate-project",
            "name": "Fixed coordinate project layout",
            "match": {"min_confidence": 1.0, "review_confidence": 0.5},
            "rack": {
                "title": {"mode": "fixed_range", "range": "G2:H2"},
                "height": {"mode": "infer_from_u_axis"},
            },
            "u_axis": {
                "direction": "descending",
                "pairing": "paired",
                "left_column": "F",
                "right_column": "I",
                "start_row": 4,
                "end_row": 11,
                "max_missing_rows": 0,
            },
            "device_area": {"mode": "fixed_range", "range": "G4:H11"},
        }
    )


def _write_profile(path: Path) -> None:
    payload = canonical_profile_payload(_fixed_coordinate_profile())
    path.write_text(
        yaml.safe_dump(payload, allow_unicode=True, sort_keys=False),
        encoding="utf-8",
    )


def _device(project, text: str):
    return next(item for item in project.devices if item.display_text == text)


def _profile_metadata(database: Path) -> dict[str, object]:
    connection = sqlite3.connect(database)
    try:
        raw = connection.execute("SELECT metadata_json FROM projects").fetchone()
        assert raw is not None
        payload = json.loads(raw[0])
        assert isinstance(payload, dict)
        return payload
    finally:
        connection.close()


def _replace_profile_metadata(database: Path, metadata: dict[str, object]) -> None:
    connection = sqlite3.connect(database)
    try:
        connection.execute(
            "UPDATE projects SET metadata_json = ?",
            (json.dumps(metadata, ensure_ascii=False, sort_keys=True),),
        )
        connection.commit()
    finally:
        connection.close()


def test_profile_drives_import_sqlite_rescan_gui_and_safe_sync(
    tmp_path: Path,
    capsys: pytest.CaptureFixture[str],
) -> None:
    workbook_path = tmp_path / "fixed-layout.xlsx"
    profile_path = tmp_path / "fixed-layout.yaml"
    database = tmp_path / "project.sqlite"
    _make_fixed_coordinate_layout(workbook_path)
    _write_profile(profile_path)
    source_before_import = workbook_path.read_bytes()

    assert analyze_workbook(workbook_path).sheets[0].racks == []
    assert (
        main(
            [
                "project",
                "import",
                str(workbook_path),
                str(database),
                "--profile",
                str(profile_path),
            ]
        )
        == 0
    )
    imported_output = json.loads(capsys.readouterr().out)
    imported = load_project(database)
    stored_profile = load_stored_profile(imported.profile_id, imported.metadata)

    assert imported_output["profile_id"] == "fixed-coordinate-project"
    assert imported.profile_id == stored_profile.profile_id == "fixed-coordinate-project"
    assert imported.metadata[PROFILE_METADATA_KEY] == make_profile_record(stored_profile)
    assert imported.racks[0].height_u == 8
    assert _device(imported, "设备 A")
    assert workbook_path.read_bytes() == source_before_import
    assert str(profile_path) not in json.dumps(imported.metadata, ensure_ascii=False)

    device_id = _device(imported, "设备 A").device_id
    rack_id = imported.racks[0].rack_id
    profile_path.write_text("this is no longer valid YAML: [", encoding="utf-8")
    workbook = load_workbook(workbook_path)
    workbook["坐标布局"]["G4"] = "设备 A-改名"
    workbook.save(workbook_path)
    workbook.close()

    session = GuiSession.open_project(database)
    rescanned = session.rescan()

    assert _device(rescanned, "设备 A-改名").device_id == device_id
    assert rescanned.racks[0].rack_id == rack_id
    assert load_stored_profile(rescanned.profile_id, rescanned.metadata) == stored_profile

    plan = session.plan_move(device_id, rack_id, 4, 5)
    assert not plan.conflicts
    result = session.apply_move()

    assert result.status == "applied"
    assert result.backup_path is not None
    persisted = load_project(database)
    placement = next(item for item in persisted.placements if item.device_id == device_id)
    mapping = next(
        item
        for item in persisted.mappings
        if item.mapping_kind == "device" and item.device_id == device_id
    )
    assert (placement.start_u, placement.end_u) == (4, 5)
    assert mapping.source_range == "G7:H8"
    assert load_stored_profile(persisted.profile_id, persisted.metadata) == stored_profile
    workbook = load_workbook(workbook_path)
    try:
        assert workbook["坐标布局"]["G7"].value == "设备 A-改名"
        assert workbook["坐标布局"]["Z1"].value == "无关内容"
    finally:
        workbook.close()


@pytest.mark.parametrize("tamper", ["missing", "hash-mismatch"])
def test_damaged_persisted_profile_fails_closed_without_touching_source(
    tmp_path: Path,
    tamper: str,
) -> None:
    workbook_path = tmp_path / "fixed-layout.xlsx"
    database = tmp_path / "project.sqlite"
    _make_fixed_coordinate_layout(workbook_path)
    project = import_project(
        workbook_path,
        database,
        profile=_fixed_coordinate_profile(),
    )
    device = _device(project, "设备 A")
    plan = plan_device_move(
        project,
        device.device_id,
        project.racks[0].rack_id,
        4,
        5,
        workbook_path=workbook_path,
    )
    assert not plan.conflicts
    source_bytes = workbook_path.read_bytes()
    metadata = _profile_metadata(database)
    if tamper == "missing":
        metadata.pop(PROFILE_METADATA_KEY)
    else:
        record = metadata[PROFILE_METADATA_KEY]
        assert isinstance(record, dict)
        payload = record["payload"]
        assert isinstance(payload, dict)
        payload["name"] = "tampered without updating the hash"
    _replace_profile_metadata(database, metadata)
    damaged_project = replace(project, metadata=metadata)
    blocked_plan = plan_device_move(
        damaged_project,
        device.device_id,
        project.racks[0].rack_id,
        4,
        5,
        workbook_path=workbook_path,
    )

    assert blocked_plan.actions == ()
    assert any(item.code == "invalid-stored-profile" for item in blocked_plan.conflicts)
    with pytest.raises(ProfileValidationError):
        load_project(database)
    with pytest.raises(ProfileValidationError):
        rescan_project(workbook_path, database)
    with pytest.raises(ProfileValidationError):
        commit_write_plan(workbook_path, database, plan)

    assert workbook_path.read_bytes() == source_bytes


def test_import_revalidates_profile_object_and_cli_rejects_forged_label(
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "fixed-layout.xlsx"
    database = tmp_path / "project.sqlite"
    _make_fixed_coordinate_layout(workbook_path)
    invalid_profile = replace(
        _fixed_coordinate_profile(),
        profile_id="FORGED PROFILE LABEL",
    )
    source_bytes = workbook_path.read_bytes()

    with pytest.raises(ProfileValidationError, match="profile_id"):
        import_workbook(workbook_path, profile=invalid_profile)
    with pytest.raises(SystemExit) as error:
        main(
            [
                "project",
                "import",
                str(workbook_path),
                str(database),
                "--profile-id",
                "forged-label",
            ]
        )

    assert error.value.code == 2
    assert not database.exists()
    assert workbook_path.read_bytes() == source_bytes

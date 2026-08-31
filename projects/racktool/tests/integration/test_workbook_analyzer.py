import json
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from racktool.cli.main import main
from racktool.core import analyze_workbook
from racktool.core.analyzer import AnalysisOptions


def _fill_descending_axis(
    sheet: Worksheet, column: int, start_row: int, height: int
) -> None:
    for offset, u_number in enumerate(range(height, 0, -1)):
        sheet.cell(start_row + offset, column, u_number)


def _fill_ascending_axis(
    sheet: Worksheet, column: int, start_row: int, height: int
) -> None:
    for offset, u_number in enumerate(range(1, height + 1)):
        sheet.cell(start_row + offset, column, u_number)


def _make_dual_axis_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "双轴"
    sheet.merge_cells("A1:C1")
    sheet["A1"] = "RACK-A"
    _fill_descending_axis(sheet, 1, 2, 12)
    _fill_descending_axis(sheet, 3, 2, 12)
    sheet.merge_cells("B2:B3")
    sheet["B2"] = "设备 A"
    sheet["B5"] = "设备 B"
    workbook.save(path)


def _make_shared_axis_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "共享轴"
    sheet.merge_cells("A1:B1")
    sheet["A1"] = "RACK-1"
    sheet.merge_cells("C1:D1")
    sheet["C1"] = "RACK-2"
    _fill_descending_axis(sheet, 1, 2, 10)
    _fill_descending_axis(sheet, 3, 2, 10)
    sheet["B2"] = "共享轴左侧设备"
    sheet.merge_cells("D2:D6")
    sheet["D2"] = "边缘机柜设备"
    workbook.save(path)


def test_analyzer_detects_dual_axis_rack_and_device_heights(tmp_path: Path) -> None:
    path = tmp_path / "dual-axis.xlsx"
    _make_dual_axis_workbook(path)

    first = analyze_workbook(path).to_dict()
    second = analyze_workbook(path).to_dict()
    analysis = first["sheets"][0]

    assert first == second
    assert len(analysis["u_axes"]) == 2
    assert len(analysis["racks"]) == 1
    assert analysis["racks"][0]["rack_name"] == "RACK-A"
    assert analysis["racks"][0]["right_axis_column"] == 3
    rack_ids = {item["candidate_id"] for item in analysis["racks"]}
    device_ids = {item["candidate_id"] for item in analysis["devices"]}
    assert len(rack_ids) == len(analysis["racks"])
    assert len(device_ids) == len(analysis["devices"])
    assert all(item["rack_candidate_id"] in rack_ids for item in analysis["placements"])
    assert all(item["device_candidate_id"] in device_ids for item in analysis["placements"])
    assert [(item["start_u"], item["end_u"]) for item in analysis["placements"]] == [
        (11, 12),
        (9, 9),
    ]


def test_analyzer_supports_shared_axis_and_single_axis_edge_rack(tmp_path: Path) -> None:
    path = tmp_path / "shared-axis.xlsx"
    _make_shared_axis_workbook(path)

    analysis = analyze_workbook(path).to_dict()["sheets"][0]

    assert [rack["rack_name"] for rack in analysis["racks"]] == ["RACK-1", "RACK-2"]
    assert analysis["racks"][0]["right_axis_column"] == 3
    assert analysis["racks"][1]["right_axis_column"] is None
    assert analysis["placements"][1]["height_u"] == 5


def test_analyzer_supports_ascending_u_axis(tmp_path: Path) -> None:
    path = tmp_path / "ascending-axis.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.merge_cells("A1:C1")
    sheet["A1"] = "RACK-UP"
    _fill_ascending_axis(sheet, 1, 2, 10)
    _fill_ascending_axis(sheet, 3, 2, 10)
    sheet.merge_cells("B2:B4")
    sheet["B2"] = "设备"
    workbook.save(path)

    analysis = analyze_workbook(path).to_dict()["sheets"][0]

    assert analysis["u_axes"][0]["direction"] == "ascending"
    assert analysis["placements"][0]["start_u"] == 1
    assert analysis["placements"][0]["end_u"] == 3


def test_analyze_cli_emits_candidate_json(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    path = tmp_path / "cli-analysis.xlsx"
    _make_dual_axis_workbook(path)

    assert main(["analyze", str(path)]) == 0
    captured = capsys.readouterr()
    payload = json.loads(captured.out)

    assert payload["sheets"][0]["racks"][0]["rack_name"] == "RACK-A"
    assert payload["sheets"][0]["placements"][0]["height_u"] == 2


def test_axis_without_merged_title_stays_unresolved(tmp_path: Path) -> None:
    path = tmp_path / "axis-only.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    _fill_descending_axis(sheet, 1, 2, 10)
    _fill_descending_axis(sheet, 3, 2, 10)
    workbook.save(path)

    analysis = analyze_workbook(path).to_dict()["sheets"][0]

    assert len(analysis["u_axes"]) == 2
    assert analysis["racks"] == []
    assert analysis["devices"] == []
    assert analysis["placements"] == []
    assert [issue["code"] for issue in analysis["issues"]] == [
        "unresolved-u-axis",
        "unresolved-u-axis",
    ]


def test_short_integer_runs_are_not_promoted_to_u_axes(tmp_path: Path) -> None:
    path = tmp_path / "short-run.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.merge_cells("A1:C1")
    sheet["A1"] = "NOT-A-RACK"
    _fill_descending_axis(sheet, 1, 2, 3)
    _fill_descending_axis(sheet, 3, 2, 3)
    workbook.save(path)

    analysis = analyze_workbook(path).to_dict()["sheets"][0]

    assert analysis["u_axes"] == []
    assert analysis["racks"] == []


def test_duplicate_rack_titles_are_reported_as_ambiguous(tmp_path: Path) -> None:
    path = tmp_path / "duplicate-titles.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    for title_row, axis_row in ((1, 2), (13, 14)):
        sheet.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=3)
        sheet.cell(title_row, 1, "RACK-DUPLICATE")
        _fill_descending_axis(sheet, 1, axis_row, 10)
        _fill_descending_axis(sheet, 3, axis_row, 10)
    workbook.save(path)

    analysis = analyze_workbook(path).to_dict()["sheets"][0]

    assert [issue["code"] for issue in analysis["issues"]] == ["duplicate-rack-title"]
    assert len(analysis["issues"][0]["candidate_ids"]) == 2


def test_overlapping_device_columns_are_reported_as_conflict(tmp_path: Path) -> None:
    path = tmp_path / "overlap.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.merge_cells("A1:D1")
    sheet["A1"] = "RACK-WIDE"
    _fill_descending_axis(sheet, 1, 2, 10)
    _fill_descending_axis(sheet, 4, 2, 10)
    sheet["B2"] = "前侧候选"
    sheet["C2"] = "后侧候选"
    workbook.save(path)

    analysis = analyze_workbook(path).to_dict()["sheets"][0]

    assert len(analysis["placements"]) == 2
    assert [issue["code"] for issue in analysis["issues"]] == [
        "overlapping-placement-candidates"
    ]
    assert analysis["issues"][0]["severity"] == "error"


def test_numeric_device_labels_are_preserved_as_text(tmp_path: Path) -> None:
    path = tmp_path / "numeric-device.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.merge_cells("A1:C1")
    sheet["A1"] = "RACK-NUMERIC"
    _fill_descending_axis(sheet, 1, 2, 10)
    _fill_descending_axis(sheet, 3, 2, 10)
    sheet["B2"] = 10001
    workbook.save(path)

    analysis = analyze_workbook(path).to_dict()["sheets"][0]

    assert analysis["devices"][0]["display_text"] == "10001"
    assert analysis["devices"][0]["value_type"] == "n"
    assert analysis["devices"][0]["confidence"] == 0.5
    assert analysis["placements"][0]["confidence"] == 0.5
    assert [issue["code"] for issue in analysis["issues"]] == [
        "non-text-device-candidate"
    ]


def test_partial_non_one_based_axis_never_forms_a_rack_or_placement(tmp_path: Path) -> None:
    path = tmp_path / "partial-axis.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.merge_cells("A1:C1")
    sheet["A1"] = "RACK-PARTIAL"
    for row, u_number in enumerate(range(20, 8, -1), start=2):
        sheet.cell(row, 1, u_number)
        sheet.cell(row, 3, u_number)
    sheet["B2"] = "设备"
    workbook.save(path)

    analysis = analyze_workbook(path).to_dict()["sheets"][0]

    assert len(analysis["u_axes"]) == 2
    assert all(axis["min_u"] == 9 for axis in analysis["u_axes"])
    assert analysis["racks"] == []
    assert analysis["devices"] == []
    assert analysis["placements"] == []
    assert [issue["code"] for issue in analysis["issues"]] == [
        "unresolved-u-axis",
        "unresolved-u-axis",
    ]


def test_one_bounded_blank_row_preserves_explicit_u_mapping(tmp_path: Path) -> None:
    path = tmp_path / "near-contiguous-axis.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.merge_cells("A1:C1")
    sheet["A1"] = "RACK-GAP"
    values = [12, 11, 10, None, 9, 8, 7, 6, 5, 4, 3, 2, 1]
    for row, u_number in enumerate(values, start=2):
        if u_number is not None:
            sheet.cell(row, 1, u_number)
            sheet.cell(row, 3, u_number)
    sheet.merge_cells("B4:B6")
    sheet["B4"] = "跨空白行设备"
    workbook.save(path)

    analysis = analyze_workbook(path).to_dict()["sheets"][0]
    rack = analysis["racks"][0]

    assert len(analysis["u_axes"]) == 2
    assert rack["height_u"] == 12
    assert rack["u_to_row"] == {
        1: 14,
        2: 13,
        3: 12,
        4: 11,
        5: 10,
        6: 9,
        7: 8,
        8: 7,
        9: 6,
        10: 4,
        11: 3,
        12: 2,
    }
    assert analysis["placements"][0]["start_u"] == 9
    assert analysis["placements"][0]["end_u"] == 10
    assert analysis["placements"][0]["height_u"] == 2


def test_more_than_one_blank_axis_row_is_not_guessed(tmp_path: Path) -> None:
    path = tmp_path / "ambiguous-axis-gap.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.merge_cells("A1:C1")
    sheet["A1"] = "RACK-AMBIGUOUS"
    values = [12, 11, 10, None, None, 9, 8, 7, 6, 5, 4, 3, 2, 1]
    for row, u_number in enumerate(values, start=2):
        if u_number is not None:
            sheet.cell(row, 1, u_number)
            sheet.cell(row, 3, u_number)
    workbook.save(path)

    analysis = analyze_workbook(path).to_dict()["sheets"][0]

    assert analysis["racks"] == []
    assert analysis["devices"] == []
    assert analysis["placements"] == []


def test_analyzer_rejects_unbounded_axis_gap_configuration(tmp_path: Path) -> None:
    path = tmp_path / "invalid-gap-configuration.xlsx"
    _make_dual_axis_workbook(path)

    with pytest.raises(ValueError, match="max_missing_rows must be 0 or 1"):
        analyze_workbook(path, AnalysisOptions(max_missing_rows=2))

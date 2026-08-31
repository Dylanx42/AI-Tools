from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from racktool.gui.session import GuiSession
from racktool.persistence import load_project


def _fill_descending_axis(
    sheet: Worksheet,
    column: int,
    start_row: int,
    height: int,
) -> None:
    for offset, u_number in enumerate(range(height, 0, -1)):
        sheet.cell(start_row + offset, column, u_number)


def _make_layout(path: Path, *, duplicate_names: bool = False) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "机柜"
    sheet.merge_cells("A1:C1")
    sheet["A1"] = "RACK-A"
    _fill_descending_axis(sheet, 1, 2, 12)
    _fill_descending_axis(sheet, 3, 2, 12)
    sheet["B2"] = "同名设备" if duplicate_names else "设备 A"
    sheet["B5"] = "同名设备" if duplicate_names else "设备 B"
    workbook.save(path)


def test_session_lists_devices_racks_mappings_and_occupancy(tmp_path: Path) -> None:
    path = tmp_path / "layout.xlsx"
    _make_layout(path)
    session = GuiSession.open_workbook(path)

    assert len(session.device_rows()) == 2
    assert session.rack_rows()[0]["rack_name"] == "RACK-A"
    assert session.mapping_rows()
    occupancy = session.occupancy_rows(session.project.racks[0].rack_id)
    assert occupancy[0]["u"] == 12
    assert any(row["display_text"] == "设备 B" for row in occupancy)


def test_session_preview_rejects_occupied_target(tmp_path: Path) -> None:
    path = tmp_path / "layout.xlsx"
    _make_layout(path)
    session = GuiSession.open_workbook(path)
    device = next(
        item for item in session.project.devices if item.display_text == "设备 B"
    )
    rack_id = session.project.racks[0].rack_id

    plan = session.plan_move(device.device_id, rack_id, 12, 12)

    assert any(item.code == "target-u-occupied" for item in plan.conflicts)
    assert session.conflict_rows()


def test_session_move_can_reopen_export_and_restore_atomically(tmp_path: Path) -> None:
    path = tmp_path / "layout.xlsx"
    _make_layout(path)
    original = path.read_bytes()
    session = GuiSession.open_workbook(path)
    device = next(
        item for item in session.project.devices if item.display_text == "设备 B"
    )
    rack_id = session.project.racks[0].rack_id

    plan = session.plan_move(device.device_id, rack_id, 4, 4)
    assert not plan.conflicts
    result = session.apply_move()
    assert result.status == "applied"
    assert session.project.source_workbook == str(path.resolve())

    reopened = GuiSession.open_project(session.database_path)
    reopened_mapping = next(
        item
        for item in reopened.project.mappings
        if item.mapping_kind == "device" and item.device_id == device.device_id
    )
    assert reopened_mapping.source_range == "B10"
    export_path = reopened.export_json(tmp_path / "project.json")
    assert export_path.is_file()

    backups = reopened.backups()
    assert backups
    reopened.restore_backup(backups[0])
    assert path.read_bytes() == original
    restored = GuiSession.open_project(reopened.database_path)
    restored_mapping = next(
        item
        for item in restored.project.mappings
        if item.mapping_kind == "device" and item.device_id == device.device_id
    )
    assert restored_mapping.source_range == "B5"


def test_session_ambiguous_rescan_keeps_memory_and_database_unchanged(
    tmp_path: Path,
) -> None:
    path = tmp_path / "layout.xlsx"
    _make_layout(path, duplicate_names=True)
    session = GuiSession.open_workbook(path)
    original_project = session.project

    workbook = load_workbook(path)
    sheet = workbook.active
    sheet["B2"] = None
    sheet["B5"] = None
    sheet["B8"] = "同名设备"
    sheet["B10"] = "同名设备"
    workbook.save(path)

    returned = session.rescan()

    assert returned == original_project
    assert session.project == original_project
    assert any(
        item.code == "ambiguous-device-identity"
        for item in session.last_rescan_conflicts
    )
    assert load_project(session.database_path).to_dict() == original_project.to_dict()


def test_session_database_failure_does_not_adopt_half_written_state(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = tmp_path / "layout.xlsx"
    _make_layout(path)
    original_workbook = path.read_bytes()
    session = GuiSession.open_workbook(path)
    original_project = session.project
    device = next(
        item for item in session.project.devices if item.display_text == "设备 B"
    )
    rack_id = session.project.racks[0].rack_id
    session.plan_move(device.device_id, rack_id, 4, 4)

    def fail_save(*args, **kwargs):
        raise OSError("forced database save failure")

    monkeypatch.setattr("racktool.core.service.save_project", fail_save)
    result = session.apply_move()

    assert result.status == "failed"
    assert session.project == original_project
    assert path.read_bytes() == original_workbook
    assert load_project(session.database_path).to_dict() == original_project.to_dict()

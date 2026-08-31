from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import yaml
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import range_boundaries
from yaml.nodes import MappingNode

from racktool.profiles.schema import (
    DeviceAreaRule,
    LayoutProfile,
    ProfileLoadError,
    ProfileMatchRule,
    ProfileValidationError,
    RackRule,
    TextRule,
    UAxisRule,
)

_PROFILE_ID = re.compile(r"^[a-z0-9]+(?:-[a-z0-9]+)*$")
_MAX_PROFILE_BYTES = 1024 * 1024
_MAX_EXCEL_COLUMN = 16_384
_MAX_EXCEL_ROW = 1_048_576


class _UniqueKeySafeLoader(yaml.SafeLoader):
    pass


def _construct_unique_mapping(
    loader: _UniqueKeySafeLoader, node: MappingNode, deep: bool = False
) -> dict[Any, Any]:
    loader.flatten_mapping(node)
    result: dict[Any, Any] = {}
    for key_node, value_node in node.value:
        key = loader.construct_object(key_node, deep=deep)
        if key in result:
            raise ProfileValidationError(f"Duplicate YAML key: {key}")
        result[key] = loader.construct_object(value_node, deep=deep)
    return result


_UniqueKeySafeLoader.add_constructor(
    yaml.resolver.BaseResolver.DEFAULT_MAPPING_TAG,
    _construct_unique_mapping,
)


def _mapping(value: Any, path: str) -> dict[str, Any]:
    if not isinstance(value, dict) or not all(isinstance(key, str) for key in value):
        raise ProfileValidationError(f"{path} must be a mapping with string keys")
    return value


def _check_keys(
    value: dict[str, Any], *, allowed: set[str], required: set[str], path: str
) -> None:
    unknown = sorted(set(value) - allowed)
    missing = sorted(required - set(value))
    if unknown:
        raise ProfileValidationError(f"{path} has unknown fields: {', '.join(unknown)}")
    if missing:
        raise ProfileValidationError(f"{path} is missing fields: {', '.join(missing)}")


def _string(value: Any, path: str) -> str:
    if not isinstance(value, str) or not value.strip():
        raise ProfileValidationError(f"{path} must be a non-empty string")
    return value.strip()


def _boolean(value: Any, path: str) -> bool:
    if not isinstance(value, bool):
        raise ProfileValidationError(f"{path} must be a boolean")
    return value


def _number(value: Any, path: str) -> float:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise ProfileValidationError(f"{path} must be a number")
    return float(value)


def _enum(value: Any, allowed: set[str], path: str) -> str:
    text = _string(value, path)
    if text not in allowed:
        raise ProfileValidationError(
            f"{path} must be one of: {', '.join(sorted(allowed))}"
        )
    return text


def _positive_integer_list(value: Any, path: str) -> tuple[int, ...]:
    if not isinstance(value, list):
        raise ProfileValidationError(f"{path} must be a list")
    items: list[int] = []
    for index, item in enumerate(value):
        if isinstance(item, bool) or not isinstance(item, int) or item <= 0:
            raise ProfileValidationError(f"{path}[{index}] must be a positive integer")
        items.append(item)
    if len(items) != len(set(items)):
        raise ProfileValidationError(f"{path} must not contain duplicates")
    return tuple(sorted(items))


def _non_negative_integer(value: Any, path: str) -> int:
    if isinstance(value, bool) or not isinstance(value, int) or value < 0:
        raise ProfileValidationError(f"{path} must be a non-negative integer")
    return int(value)


def _positive_integer(value: Any, path: str) -> int:
    if isinstance(value, bool) or not isinstance(value, int) or value < 1:
        raise ProfileValidationError(f"{path} must be a positive integer")
    return int(value)


def _column(value: Any, path: str) -> int:
    if isinstance(value, bool):
        raise ProfileValidationError(f"{path} must be a positive column number or letter")
    if isinstance(value, int):
        if not 1 <= value <= _MAX_EXCEL_COLUMN:
            raise ProfileValidationError(f"{path} must be a positive column number or letter")
        return value
    text = _string(value, path).upper()
    try:
        column = column_index_from_string(text)
    except ValueError as error:
        raise ProfileValidationError(
            f"{path} must be a positive column number or letter"
        ) from error
    if column > _MAX_EXCEL_COLUMN:
        raise ProfileValidationError(f"{path} exceeds the XLSX column limit")
    return column


def _range(value: Any, path: str) -> str:
    text = _string(value, path).upper().replace("$", "")
    try:
        min_col, min_row, max_col, max_row = range_boundaries(text)
    except ValueError as error:
        raise ProfileValidationError(f"{path} must be a valid A1 range") from error
    if min_col is None or min_row is None or max_col is None or max_row is None:
        raise ProfileValidationError(f"{path} must be a bounded A1 range")
    if min_col > max_col or min_row > max_row:
        raise ProfileValidationError(f"{path} must not be inverted")
    if max_col > _MAX_EXCEL_COLUMN or max_row > _MAX_EXCEL_ROW:
        raise ProfileValidationError(f"{path} exceeds XLSX worksheet limits")
    start = f"{get_column_letter(min_col)}{min_row}"
    end = f"{get_column_letter(max_col)}{max_row}"
    return start if start == end else f"{start}:{end}"


def _string_list(value: Any, path: str) -> tuple[str, ...]:
    if not isinstance(value, list):
        raise ProfileValidationError(f"{path} must be a list")
    items = tuple(_string(item, f"{path}[{index}]") for index, item in enumerate(value))
    if len(items) != len(set(items)):
        raise ProfileValidationError(f"{path} must not contain duplicates")
    return items


def validate_profile_data(raw: Any) -> LayoutProfile:
    root = _mapping(raw, "profile")
    _check_keys(
        root,
        allowed={
            "schema_version",
            "profile_id",
            "name",
            "match",
            "rack",
            "u_axis",
            "device_area",
            "text",
        },
        required={
            "schema_version",
            "profile_id",
            "name",
            "match",
            "rack",
            "u_axis",
            "device_area",
        },
        path="profile",
    )

    schema_version = root["schema_version"]
    if isinstance(schema_version, bool) or schema_version != 1:
        raise ProfileValidationError("profile.schema_version must be the integer 1")
    profile_id = _string(root["profile_id"], "profile.profile_id")
    if _PROFILE_ID.fullmatch(profile_id) is None:
        raise ProfileValidationError(
            "profile.profile_id must use lowercase ASCII words separated by hyphens"
        )

    match_raw = _mapping(root["match"], "profile.match")
    _check_keys(
        match_raw,
        allowed={
            "sheet_name_regex",
            "min_confidence",
            "review_confidence",
            "allow_multiple_sheets",
            "require_issue_free",
        },
        required=set(),
        path="profile.match",
    )
    sheet_name_regex = match_raw.get("sheet_name_regex")
    if sheet_name_regex is not None:
        sheet_name_regex = _string(sheet_name_regex, "profile.match.sheet_name_regex")
        try:
            re.compile(sheet_name_regex)
        except re.error as error:
            raise ProfileValidationError(
                f"profile.match.sheet_name_regex is invalid: {error}"
            ) from error
    min_confidence = _number(match_raw.get("min_confidence", 1.0), "profile.match.min_confidence")
    review_confidence = _number(
        match_raw.get("review_confidence", 0.5),
        "profile.match.review_confidence",
    )
    if not 0 < min_confidence <= 1:
        raise ProfileValidationError("profile.match.min_confidence must be in (0, 1]")
    if not 0 <= review_confidence < min_confidence:
        raise ProfileValidationError(
            "profile.match.review_confidence must be in [0, min_confidence)"
        )

    rack_raw = _mapping(root["rack"], "profile.rack")
    _check_keys(
        rack_raw,
        allowed={"title", "height"},
        required={"title", "height"},
        path="profile.rack",
    )
    title_raw = _mapping(rack_raw["title"], "profile.rack.title")
    height_raw = _mapping(rack_raw["height"], "profile.rack.height")
    _check_keys(
        title_raw,
        allowed={"mode", "range"},
        required={"mode"},
        path="profile.rack.title",
    )
    _check_keys(
        height_raw,
        allowed={"mode"},
        required={"mode"},
        path="profile.rack.height",
    )

    title_mode = _enum(
        title_raw["mode"],
        {"merged_cell_above_u_axis", "fixed_range"},
        "profile.rack.title.mode",
    )
    title_range = (
        _range(title_raw["range"], "profile.rack.title.range")
        if "range" in title_raw
        else None
    )
    if title_mode == "fixed_range" and title_range is None:
        raise ProfileValidationError("profile.rack.title.range is required for fixed_range")
    if title_mode != "fixed_range" and title_range is not None:
        raise ProfileValidationError(
            "profile.rack.title.range is only valid for fixed_range"
        )

    axis_raw = _mapping(root["u_axis"], "profile.u_axis")
    _check_keys(
        axis_raw,
        allowed={
            "direction",
            "pairing",
            "allowed_heights",
            "left_column",
            "right_column",
            "start_row",
            "end_row",
            "max_missing_rows",
        },
        required=set(),
        path="profile.u_axis",
    )
    direction = _enum(
        axis_raw.get("direction", "any"),
        {"ascending", "descending", "any"},
        "profile.u_axis.direction",
    )
    pairing = _enum(
        axis_raw.get("pairing", "any"),
        {"paired", "single_axis_edge", "mixed", "any"},
        "profile.u_axis.pairing",
    )
    left_column = (
        _column(axis_raw["left_column"], "profile.u_axis.left_column")
        if "left_column" in axis_raw
        else None
    )
    right_column = (
        _column(axis_raw["right_column"], "profile.u_axis.right_column")
        if "right_column" in axis_raw
        else None
    )
    start_row = (
        _positive_integer(axis_raw["start_row"], "profile.u_axis.start_row")
        if "start_row" in axis_raw
        else None
    )
    end_row = (
        _positive_integer(axis_raw["end_row"], "profile.u_axis.end_row")
        if "end_row" in axis_raw
        else None
    )
    max_missing_rows = _non_negative_integer(
        axis_raw.get("max_missing_rows", 1),
        "profile.u_axis.max_missing_rows",
    )
    if max_missing_rows > 1:
        raise ProfileValidationError("profile.u_axis.max_missing_rows must be 0 or 1")
    coordinates = (left_column, start_row, end_row)
    if any(value is not None for value in coordinates) and any(
        value is None for value in coordinates
    ):
        raise ProfileValidationError(
            "profile.u_axis left_column, start_row, and end_row must be provided together"
        )
    if start_row is not None and end_row is not None and end_row < start_row:
        raise ProfileValidationError("profile.u_axis.end_row must not precede start_row")
    if right_column is not None and left_column is None:
        raise ProfileValidationError(
            "profile.u_axis.right_column requires left_column coordinates"
        )
    if left_column is not None and right_column is not None and right_column <= left_column:
        raise ProfileValidationError("profile.u_axis.right_column must be right of left_column")

    device_raw = _mapping(root["device_area"], "profile.device_area")
    _check_keys(
        device_raw,
        allowed={"mode", "range"},
        required={"mode"},
        path="profile.device_area",
    )
    device_mode = _enum(
        device_raw["mode"],
        {"between_u_axes", "between_or_edge", "fixed_range"},
        "profile.device_area.mode",
    )
    device_range = (
        _range(device_raw["range"], "profile.device_area.range")
        if "range" in device_raw
        else None
    )
    if device_mode == "fixed_range" and device_range is None:
        raise ProfileValidationError("profile.device_area.range is required for fixed_range")
    if device_mode != "fixed_range" and device_range is not None:
        raise ProfileValidationError(
            "profile.device_area.range is only valid for fixed_range"
        )
    if pairing == "single_axis_edge" and device_mode == "between_u_axes":
        raise ProfileValidationError(
            "single_axis_edge pairing conflicts with between_u_axes device area"
        )
    if pairing == "mixed" and device_mode != "between_or_edge":
        raise ProfileValidationError(
            "mixed pairing requires between_or_edge device area"
        )
    coordinate_fallback = title_mode == "fixed_range" or device_mode == "fixed_range"
    if not coordinate_fallback and any(
        value is not None
        for value in (left_column, right_column, start_row, end_row)
    ):
        raise ProfileValidationError(
            "profile.u_axis coordinates require fixed coordinate fallback"
        )
    if coordinate_fallback:
        if title_mode != "fixed_range" or device_mode != "fixed_range":
            raise ProfileValidationError(
                "fixed coordinate fallback requires fixed_range title and device area"
            )
        if left_column is None or start_row is None or end_row is None:
            raise ProfileValidationError(
                "fixed coordinate fallback requires U-axis column and row coordinates"
            )
        if direction == "any":
            raise ProfileValidationError(
                "fixed coordinate fallback requires an explicit U-axis direction"
            )
        if pairing == "paired" and right_column is None:
            raise ProfileValidationError(
                "paired fixed coordinate fallback requires right_column"
            )
        if pairing == "single_axis_edge" and right_column is not None:
            raise ProfileValidationError(
                "single_axis_edge fixed coordinate fallback must not define right_column"
            )
        if pairing not in {"paired", "single_axis_edge"}:
            raise ProfileValidationError(
                "fixed coordinate fallback requires paired or single_axis_edge pairing"
            )
        if device_range is None:
            raise ProfileValidationError("fixed coordinate fallback requires device range")
        device_min_col, device_min_row, device_max_col, device_max_row = range_boundaries(
            device_range
        )
        if (
            device_min_col is None
            or device_min_row is None
            or device_max_col is None
            or device_max_row is None
        ):
            raise ProfileValidationError("profile.device_area.range must be bounded")
        if device_min_row < start_row or device_max_row > end_row:
            raise ProfileValidationError(
                "profile.device_area.range rows must stay within the U-axis rows"
            )
        if device_min_col <= left_column:
            raise ProfileValidationError(
                "profile.device_area.range must be right of the left U-axis"
            )
        if right_column is not None and device_max_col >= right_column:
            raise ProfileValidationError(
                "profile.device_area.range must stay between paired U axes"
            )

    text_raw = _mapping(root.get("text", {}), "profile.text")
    _check_keys(
        text_raw,
        allowed={"ignore_exact"},
        required=set(),
        path="profile.text",
    )

    return LayoutProfile(
        schema_version=1,
        profile_id=profile_id,
        name=_string(root["name"], "profile.name"),
        match=ProfileMatchRule(
            sheet_name_regex=sheet_name_regex,
            min_confidence=min_confidence,
            review_confidence=review_confidence,
            allow_multiple_sheets=_boolean(
                match_raw.get("allow_multiple_sheets", False),
                "profile.match.allow_multiple_sheets",
            ),
            require_issue_free=_boolean(
                match_raw.get("require_issue_free", True),
                "profile.match.require_issue_free",
            ),
        ),
        rack=RackRule(
            title_mode=title_mode,  # type: ignore[arg-type]
            height_mode=_enum(
                height_raw["mode"],
                {"infer_from_u_axis"},
                "profile.rack.height.mode",
            ),  # type: ignore[arg-type]
            title_range=title_range,
        ),
        u_axis=UAxisRule(
            direction=direction,  # type: ignore[arg-type]
            pairing=pairing,  # type: ignore[arg-type]
            allowed_heights=_positive_integer_list(
                axis_raw.get("allowed_heights", []),
                "profile.u_axis.allowed_heights",
            ),
            left_column=left_column,
            right_column=right_column,
            start_row=start_row,
            end_row=end_row,
            max_missing_rows=max_missing_rows,
        ),
        device_area=DeviceAreaRule(
            mode=device_mode,  # type: ignore[arg-type]
            source_range=device_range,
        ),
        text=TextRule(
            ignore_exact=_string_list(
                text_raw.get("ignore_exact", []),
                "profile.text.ignore_exact",
            )
        ),
    )


def load_profile(path: Path) -> LayoutProfile:
    profile_path = path.expanduser()
    if profile_path.suffix.lower() not in {".yaml", ".yml"}:
        raise ProfileLoadError("Profile files must use .yaml or .yml")
    if not profile_path.is_file():
        raise ProfileLoadError(f"Profile file is missing: {profile_path}")
    if profile_path.stat().st_size > _MAX_PROFILE_BYTES:
        raise ProfileLoadError("Profile file exceeds the 1 MiB safety limit")
    try:
        text = profile_path.read_text(encoding="utf-8")
    except UnicodeError as error:
        raise ProfileLoadError("Profile file must be UTF-8") from error
    try:
        raw = yaml.load(text, Loader=_UniqueKeySafeLoader)
    except ProfileValidationError:
        raise
    except yaml.YAMLError as error:
        raise ProfileLoadError(f"Invalid YAML: {error}") from error
    return validate_profile_data(raw)

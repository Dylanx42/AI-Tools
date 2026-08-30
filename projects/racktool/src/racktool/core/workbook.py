from __future__ import annotations

import hashlib
import json
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

from racktool.models.domain import CellRange
from racktool.models.workbook import CellInfo, SheetInfo, WorkbookInfo


def _color(color: Color | None) -> dict[str, Any] | None:
    if color is None:
        return None
    color_type = color.type
    value = getattr(color, color_type, None) if color_type is not None else None
    return {
        "type": color_type,
        "value": value,
        "tint": color.tint,
    }


def _style_signature(cell: Cell) -> str:
    font = cell.font
    fill = cell.fill
    border = cell.border
    alignment = cell.alignment
    payload = {
        "alignment": {
            "horizontal": alignment.horizontal,
            "vertical": alignment.vertical,
            "wrap_text": alignment.wrap_text,
            "text_rotation": alignment.text_rotation,
        },
        "border": {
            side: {
                "style": getattr(border, side).style,
                "color": _color(getattr(border, side).color),
            }
            for side in ("left", "right", "top", "bottom")
        },
        "fill": {
            "fill_type": fill.fill_type,
            "fg_color": _color(fill.fgColor),
            "bg_color": _color(fill.bgColor),
        },
        "font": {
            "name": font.name,
            "size": font.sz,
            "bold": font.b,
            "italic": font.i,
            "underline": font.u,
            "color": _color(font.color),
        },
        "number_format": cell.number_format,
        "protection": {"locked": cell.protection.locked, "hidden": cell.protection.hidden},
    }
    encoded = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(encoded.encode("utf-8")).hexdigest()[:16]


def _json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def _used_range(cells: list[Cell]) -> str | None:
    if not cells:
        return None
    rows = [cell.row for cell in cells]
    columns = [cell.column for cell in cells]
    min_row, max_row = min(rows), max(rows)
    min_col, max_col = min(columns), max(columns)
    a1 = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
    return CellRange(min_row, max_row, min_col, max_col, a1).a1


def _scan_sheet(sheet: Worksheet, index: int) -> SheetInfo:
    non_empty = [
        cell
        for row in sheet.iter_rows()
        for cell in row
        if isinstance(cell, Cell) and cell.value is not None
    ]
    cells = [
        CellInfo(cell.coordinate, _json_value(cell.value), cell.data_type, _style_signature(cell))
        for cell in non_empty
    ]
    row_heights = {
        str(row): float(dimension.height)
        for row, dimension in sorted(sheet.row_dimensions.items())
        if dimension.height is not None
    }
    column_widths = {
        column: float(dimension.width)
        for column, dimension in sorted(sheet.column_dimensions.items())
        if dimension.width is not None
    }
    return SheetInfo(
        name=sheet.title,
        index=index,
        state=sheet.sheet_state,
        reported_dimension=sheet.calculate_dimension(),
        used_range=_used_range(non_empty),
        cells=cells,
        merged_ranges=sorted(str(cell_range) for cell_range in sheet.merged_cells.ranges),
        row_heights=row_heights,
        column_widths=column_widths,
        default_row_height=sheet.sheet_format.defaultRowHeight,
        default_column_width=sheet.sheet_format.defaultColWidth,
    )


def scan_workbook(path: Path) -> WorkbookInfo:
    """Read XLSX structure without mutating the source workbook."""
    workbook_path = path.expanduser()
    if workbook_path.suffix.lower() != ".xlsx":
        raise ValueError("RackTool inspect currently supports only .xlsx files")
    if not workbook_path.is_file():
        raise FileNotFoundError(workbook_path)

    try:
        workbook = load_workbook(workbook_path, read_only=False, data_only=False)
    except (BadZipFile, InvalidFileException) as error:
        raise ValueError(f"Invalid XLSX workbook: {workbook_path}") from error
    try:
        sheets = [_scan_sheet(sheet, index) for index, sheet in enumerate(workbook.worksheets)]
        return WorkbookInfo(format="xlsx", sheets=sheets)
    finally:
        workbook.close()

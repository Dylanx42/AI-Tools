from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass, replace
from pathlib import Path
from typing import Literal

from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

from racktool.core.analyzer import analyze_workbook
from racktool.core.identity import IdFactory, default_id_factory, normalize_path, sha256_file
from racktool.models.analysis import (
    DeviceCandidate,
    PlacementCandidate,
    RackCandidate,
    WorkbookAnalysis,
)
from racktool.models.domain import CellRange, Device, Placement, Rack
from racktool.models.project import (
    IdentityConflict,
    MappingKind,
    RackProject,
    RescanResult,
    SourceMapping,
)
from racktool.profiles.fingerprint import fingerprint_workbook
from racktool.profiles.schema import LayoutProfile, ProfileError
from racktool.profiles.storage import (
    load_stored_profile,
    metadata_with_profile,
    normalize_profile,
)


@dataclass(frozen=True, slots=True)
class _RackObservation:
    sheet_name: str
    candidate: RackCandidate


@dataclass(frozen=True, slots=True)
class _DeviceObservation:
    sheet_name: str
    device: DeviceCandidate
    placement: PlacementCandidate
    rack_candidate_id: str


def _cell_range(a1: str) -> CellRange:
    min_col, min_row, max_col, max_row = range_boundaries(a1)
    if min_col is None or min_row is None or max_col is None or max_row is None:
        raise ValueError(f"Invalid cell range: {a1}")
    return CellRange(min_row, max_row, min_col, max_col, a1)


def _a1(min_row: int, max_row: int, min_col: int, max_col: int) -> str:
    start = f"{get_column_letter(min_col)}{min_row}"
    end = f"{get_column_letter(max_col)}{max_row}"
    return start if start == end else f"{start}:{end}"


def _rack_bounds(rack: RackCandidate) -> CellRange:
    title = _cell_range(rack.title_range)
    right = rack.right_axis_column or max(rack.device_columns + [rack.left_axis_column])
    min_row = min(title.min_row, rack.start_row)
    max_row = max(title.max_row, rack.end_row)
    min_col = min(title.min_col, rack.left_axis_column)
    max_col = max(title.max_col, right)
    return CellRange(
        min_row,
        max_row,
        min_col,
        max_col,
        _a1(min_row, max_row, min_col, max_col),
    )


def _conflict(
    code: str,
    message: str,
    *,
    severity: Literal["warning", "error"] = "error",
    entity_ids: list[str] | None = None,
    candidate_refs: list[str] | None = None,
    evidence: list[str] | None = None,
) -> IdentityConflict:
    return IdentityConflict(
        code=code,
        severity=severity,
        message=message,
        entity_ids=entity_ids or [],
        candidate_refs=candidate_refs or [],
        evidence=evidence or [],
    )


def _dedupe_conflicts(items: list[IdentityConflict]) -> list[IdentityConflict]:
    seen: set[tuple[object, ...]] = set()
    result: list[IdentityConflict] = []
    for item in items:
        key = (
            item.code,
            item.severity,
            item.message,
            tuple(item.entity_ids),
            tuple(item.candidate_refs),
            tuple(item.evidence),
        )
        if key not in seen:
            seen.add(key)
            result.append(item)
    return result


def _analysis_conflicts(analysis: WorkbookAnalysis) -> list[IdentityConflict]:
    conflicts: list[IdentityConflict] = []
    for sheet in analysis.sheets:
        for issue in sheet.issues:
            if issue.severity == "info":
                continue
            conflicts.append(
                _conflict(
                    issue.code,
                    issue.message,
                    severity=issue.severity,
                    candidate_refs=list(issue.candidate_ids),
                    evidence=[f"{sheet.name}!{item}" for item in issue.source_ranges],
                )
            )
    return conflicts


def _observations(
    analysis: WorkbookAnalysis,
) -> tuple[
    list[_RackObservation],
    list[_DeviceObservation],
    list[IdentityConflict],
]:
    racks: list[_RackObservation] = []
    devices: list[_DeviceObservation] = []
    conflicts: list[IdentityConflict] = []
    for sheet in analysis.sheets:
        rack_ids = {item.candidate_id for item in sheet.racks}
        device_by_id = {item.candidate_id: item for item in sheet.devices}
        placement_by_device: dict[str, PlacementCandidate] = {}
        for sheet_placement in sheet.placements:
            if sheet_placement.device_candidate_id in placement_by_device:
                conflicts.append(
                    _conflict(
                        "ambiguous-device-placement",
                        "A device candidate has multiple placements",
                        candidate_refs=[sheet_placement.device_candidate_id],
                    )
                )
                continue
            placement_by_device[sheet_placement.device_candidate_id] = sheet_placement
        racks.extend(_RackObservation(sheet.name, item) for item in sheet.racks)
        for device_id, device in device_by_id.items():
            candidate_placement = placement_by_device.get(device_id)
            if candidate_placement is None:
                conflicts.append(
                    _conflict(
                        "unplaced-device-candidate",
                        "A device candidate has no rack placement",
                        candidate_refs=[device_id],
                    )
                )
                continue
            if candidate_placement.rack_candidate_id not in rack_ids:
                conflicts.append(
                    _conflict(
                        "unmapped-device-rack",
                        "A device placement references an unknown rack candidate",
                        candidate_refs=[device_id, candidate_placement.rack_candidate_id],
                    )
                )
                continue
            devices.append(
                _DeviceObservation(
                    sheet.name,
                    device,
                    candidate_placement,
                    candidate_placement.rack_candidate_id,
                )
            )
    return racks, devices, conflicts


def _make_rack(
    rack: RackCandidate,
    rack_id: str,
    sheet_name: str,
    profile_id: str | None,
) -> Rack:
    return Rack(
        rack_id=rack_id,
        rack_name=rack.rack_name,
        height_u=rack.height_u,
        source_sheet=sheet_name,
        bounds=_rack_bounds(rack),
        profile_id=profile_id,
        confidence=rack.confidence,
        start_row=rack.start_row,
        end_row=rack.end_row,
        left_axis_column=rack.left_axis_column,
        right_axis_column=rack.right_axis_column,
        device_columns=list(rack.device_columns),
        direction=rack.direction,
        u_to_row=dict(rack.u_to_row),
        title_range=rack.title_range,
        status="active",
    )


def _make_device(
    device: DeviceCandidate,
    device_id: str,
    existing: Device | None = None,
) -> Device:
    return Device(
        device_id=device_id,
        display_text=device.display_text,
        canonical_name=existing.canonical_name if existing is not None else None,
        attributes=dict(existing.attributes) if existing is not None else {},
        confidence=device.confidence,
        style_signature=device.style_signature,
    )


def _make_mapping(
    mapping_id: str,
    fingerprint: str,
    sheet_name: str,
    source_range: str,
    mapping_kind: MappingKind,
    *,
    device_id: str | None = None,
    rack_id: str | None = None,
    style_signature: str | None = None,
    confidence: float | None = None,
) -> SourceMapping:
    return SourceMapping(
        mapping_id=mapping_id,
        workbook_fingerprint=fingerprint,
        sheet_name=sheet_name,
        source_range=source_range,
        mapping_kind=mapping_kind,
        device_id=device_id,
        rack_id=rack_id,
        style_signature=style_signature,
        confidence=confidence,
    )


def validate_project(project: RackProject) -> list[IdentityConflict]:
    conflicts: list[IdentityConflict] = []
    try:
        load_stored_profile(project.profile_id, project.metadata)
    except ProfileError as error:
        conflicts.append(
            _conflict(
                "invalid-stored-profile",
                "Project Profile semantics are missing, damaged, or inconsistent",
                evidence=[str(error)],
            )
        )
    racks_by_id = {item.rack_id: item for item in project.racks}
    devices_by_id = {item.device_id: item for item in project.devices}
    if len(racks_by_id) != len(project.racks):
        conflicts.append(_conflict("duplicate-rack-id", "Rack identities are not unique"))
    if len(devices_by_id) != len(project.devices):
        conflicts.append(_conflict("duplicate-device-id", "Device identities are not unique"))
    if len({item.placement_id for item in project.placements}) != len(project.placements):
        conflicts.append(_conflict("duplicate-placement-id", "Placement identities are not unique"))
    if len({item.mapping_id for item in project.mappings}) != len(project.mappings):
        conflicts.append(_conflict("duplicate-mapping-id", "Mapping identities are not unique"))

    active_placements: dict[str, list[Placement]] = defaultdict(list)
    occupied: dict[tuple[str, int], list[str]] = defaultdict(list)
    for placement in project.placements:
        rack = racks_by_id.get(placement.rack_id)
        if placement.device_id not in devices_by_id or rack is None:
            conflicts.append(
                _conflict(
                    "dangling-placement-reference",
                    "A Placement references an unknown device or rack",
                    entity_ids=[placement.placement_id],
                )
            )
            continue
        if placement.status == "active":
            active_placements[placement.device_id].append(placement)
            if rack.status != "active":
                conflicts.append(
                    _conflict(
                        "active-placement-on-missing-rack",
                        "An active Placement references a missing rack",
                        entity_ids=[placement.placement_id, rack.rack_id],
                    )
                )
            units = range(
                min(placement.start_u, placement.end_u),
                max(placement.start_u, placement.end_u) + 1,
            )
            for unit in units:
                if unit < 1 or unit > rack.height_u or unit not in rack.u_to_row:
                    conflicts.append(
                        _conflict(
                            "placement-out-of-rack-bounds",
                            "An active Placement is outside its rack U axis",
                            entity_ids=[placement.placement_id, rack.rack_id],
                        )
                    )
                    break
                occupied[(rack.rack_id, unit)].append(placement.placement_id)

    overlapping = sorted(
        {
            placement_id
            for placement_ids in occupied.values()
            if len(placement_ids) > 1
            for placement_id in placement_ids
        }
    )
    if overlapping:
        conflicts.append(
            _conflict(
                "overlapping-active-placements",
                "Active placements overlap",
                entity_ids=overlapping,
            )
        )

    device_mappings: dict[str, list[SourceMapping]] = defaultdict(list)
    rack_mappings: dict[str, list[SourceMapping]] = defaultdict(list)
    for mapping in project.mappings:
        if mapping.device_id is not None and mapping.device_id not in devices_by_id:
            conflicts.append(
                _conflict(
                    "dangling-device-mapping",
                    "A Mapping references an unknown device",
                    entity_ids=[mapping.mapping_id, mapping.device_id],
                )
            )
        if mapping.rack_id is not None and mapping.rack_id not in racks_by_id:
            conflicts.append(
                _conflict(
                    "dangling-rack-mapping",
                    "A Mapping references an unknown rack",
                    entity_ids=[mapping.mapping_id, mapping.rack_id],
                )
            )
        if mapping.mapping_kind == "device" and mapping.device_id is not None:
            device_mappings[mapping.device_id].append(mapping)
        if mapping.mapping_kind == "rack_title" and mapping.rack_id is not None:
            rack_mappings[mapping.rack_id].append(mapping)

    for rack in project.racks:
        mappings = rack_mappings.get(rack.rack_id, [])
        if rack.status == "active":
            current_mappings = [
                item
                for item in mappings
                if item.workbook_fingerprint == project.workbook_fingerprint
            ]
            if len(current_mappings) != 1:
                conflicts.append(
                    _conflict(
                        "ambiguous-rack-mapping",
                        "An active rack must have exactly one current source Mapping",
                        entity_ids=[
                            rack.rack_id,
                            *[item.mapping_id for item in current_mappings],
                        ],
                    )
                )
            else:
                mapping = current_mappings[0]
                if (
                    mapping.sheet_name != rack.source_sheet
                    or mapping.source_range != rack.title_range
                ):
                    conflicts.append(
                        _conflict(
                            "rack-mapping-source-mismatch",
                            "Rack identity and source Mapping disagree on title location",
                            entity_ids=[rack.rack_id, mapping.mapping_id],
                        )
                    )

    for device in project.devices:
        placements = active_placements.get(device.device_id, [])
        mappings = device_mappings.get(device.device_id, [])
        if len(placements) > 1:
            conflicts.append(
                _conflict(
                    "ambiguous-active-placement",
                    "A device has multiple active Placements",
                    entity_ids=[device.device_id, *[item.placement_id for item in placements]],
                )
            )
        if len(placements) == 1:
            placement = placements[0]
            active_mappings = [
                item
                for item in mappings
                if item.workbook_fingerprint == project.workbook_fingerprint
            ]
            if len(active_mappings) != 1:
                conflicts.append(
                    _conflict(
                        "ambiguous-device-mapping",
                        "An active device must have exactly one current source Mapping",
                        entity_ids=[
                            device.device_id,
                            *[item.mapping_id for item in active_mappings],
                        ],
                    )
                )
                continue
            mapping = active_mappings[0]
            if mapping.rack_id != placement.rack_id:
                conflicts.append(
                    _conflict(
                        "mapping-placement-rack-mismatch",
                        "Device Mapping and Placement disagree on rack identity",
                        entity_ids=[
                            device.device_id,
                            placement.rack_id,
                            mapping.rack_id or "",
                        ],
                    )
                )
            rack = racks_by_id.get(placement.rack_id)
            if rack is not None:
                try:
                    source = _cell_range(mapping.source_range)
                except ValueError:
                    conflicts.append(
                        _conflict(
                            "invalid-device-source-range",
                            "Device Mapping has an invalid source range",
                            entity_ids=[mapping.mapping_id],
                        )
                    )
                else:
                    if mapping.sheet_name != rack.source_sheet:
                        conflicts.append(
                            _conflict(
                                "mapping-sheet-rack-mismatch",
                                "Device Mapping and rack disagree on source Sheet",
                                entity_ids=[mapping.mapping_id, rack.rack_id],
                            )
                        )
                    if not set(range(source.min_col, source.max_col + 1)).issubset(
                        set(rack.device_columns)
                    ):
                        conflicts.append(
                            _conflict(
                                "mapping-outside-rack-columns",
                                "Device Mapping is outside its rack device columns",
                                entity_ids=[mapping.mapping_id, rack.rack_id],
                            )
                        )
                    expected_rows = [
                        rack.u_to_row[unit]
                        for unit in range(
                            min(placement.start_u, placement.end_u),
                            max(placement.start_u, placement.end_u) + 1,
                        )
                        if unit in rack.u_to_row
                    ]
                    if (
                        not expected_rows
                        or source.min_row != min(expected_rows)
                        or source.max_row != max(expected_rows)
                    ):
                        conflicts.append(
                            _conflict(
                                "mapping-placement-range-mismatch",
                                "Device Mapping range and Placement U coordinates disagree",
                                entity_ids=[
                                    mapping.mapping_id,
                                    placement.placement_id,
                                ],
                            )
                        )
    return _dedupe_conflicts(conflicts)


def project_error_conflicts(project: RackProject) -> list[IdentityConflict]:
    return _dedupe_conflicts(
        [item for item in project.conflicts if item.severity == "error"]
        + [item for item in validate_project(project) if item.severity == "error"]
    )


def _with_validation(
    project: RackProject,
    base_conflicts: list[IdentityConflict],
) -> RackProject:
    return replace(
        project,
        conflicts=_dedupe_conflicts(base_conflicts + validate_project(project)),
    )


def _build_from_analysis(
    workbook_path: Path,
    analysis: WorkbookAnalysis,
    *,
    project_id: str,
    profile: LayoutProfile | None,
    id_factory: IdFactory,
) -> RackProject:
    source = normalize_path(workbook_path)
    normalized_profile = normalize_profile(profile) if profile is not None else None
    profile_id = normalized_profile.profile_id if normalized_profile is not None else None
    fingerprint = fingerprint_workbook(source, analysis)
    rack_observations, device_observations, observation_conflicts = _observations(analysis)
    rack_ids = {
        item.candidate.candidate_id: id_factory("RACK") for item in rack_observations
    }
    device_ids = {
        item.device.candidate_id: id_factory("DEV") for item in device_observations
    }
    racks = [
        _make_rack(item.candidate, rack_ids[item.candidate.candidate_id], item.sheet_name, profile_id)
        for item in rack_observations
    ]
    devices = [
        _make_device(item.device, device_ids[item.device.candidate_id])
        for item in device_observations
    ]
    placements = [
        Placement(
            placement_id=id_factory("PLC"),
            device_id=device_ids[item.device.candidate_id],
            rack_id=rack_ids[item.rack_candidate_id],
            start_u=item.placement.start_u,
            end_u=item.placement.end_u,
            status="active",
        )
        for item in device_observations
    ]
    mappings: list[SourceMapping] = []
    for rack_observation in rack_observations:
        mappings.append(
            _make_mapping(
                id_factory("MAP"),
                fingerprint.workbook_sha256,
                rack_observation.sheet_name,
                rack_observation.candidate.title_range,
                "rack_title",
                rack_id=rack_ids[rack_observation.candidate.candidate_id],
                confidence=rack_observation.candidate.confidence,
            )
        )
    for device_observation in device_observations:
        mappings.append(
            _make_mapping(
                id_factory("MAP"),
                fingerprint.workbook_sha256,
                device_observation.sheet_name,
                device_observation.device.source_range,
                "device",
                device_id=device_ids[device_observation.device.candidate_id],
                rack_id=rack_ids[device_observation.rack_candidate_id],
                confidence=device_observation.device.confidence,
                style_signature=device_observation.device.style_signature,
            )
        )
    project = RackProject(
        project_id=project_id,
        source_workbook=str(source),
        workbook_fingerprint=fingerprint.workbook_sha256,
        layout_fingerprint=fingerprint.layout_sha256,
        profile_id=profile_id,
        racks=racks,
        devices=devices,
        placements=placements,
        mappings=mappings,
        conflicts=[],
        metadata=metadata_with_profile(
            {"schema_version": 1},
            normalized_profile,
        ),
    )
    return _with_validation(
        project,
        _analysis_conflicts(analysis) + observation_conflicts,
    )


def _analyze_workbook_with_profile(
    workbook_path: Path,
    profile: LayoutProfile,
) -> WorkbookAnalysis:
    # Import the concrete implementation only after both public packages have
    # initialized. Importing through racktool.profiles here creates a package
    # initialization cycle when profiles.apply first imports core.analyzer.
    from racktool.profiles.apply import analyze_profiled_workbook

    return analyze_profiled_workbook(workbook_path, profile)


def import_workbook(
    workbook_path: Path,
    *,
    profile: LayoutProfile | None = None,
    id_factory: IdFactory | None = None,
) -> RackProject:
    source = normalize_path(workbook_path)
    factory = id_factory or default_id_factory
    normalized_profile = normalize_profile(profile) if profile is not None else None
    analysis = (
        _analyze_workbook_with_profile(source, normalized_profile)
        if normalized_profile is not None
        else analyze_workbook(source)
    )
    return _build_from_analysis(
        source,
        analysis,
        project_id=factory("PRJ"),
        profile=normalized_profile,
        id_factory=factory,
    )


def _entity_mappings(
    project: RackProject,
    kind: MappingKind,
) -> dict[str, SourceMapping]:
    grouped: dict[str, list[SourceMapping]] = defaultdict(list)
    for mapping in project.mappings:
        if mapping.mapping_kind != kind:
            continue
        entity_id = mapping.rack_id if kind == "rack_title" else mapping.device_id
        if entity_id is not None:
            grouped[entity_id].append(mapping)
    return {
        entity_id: mappings[0]
        for entity_id, mappings in grouped.items()
        if len(mappings) == 1
    }


def _range_mappings(
    project: RackProject,
    kind: MappingKind,
) -> dict[tuple[str, str], str]:
    grouped: dict[tuple[str, str], list[str]] = defaultdict(list)
    for mapping in project.mappings:
        if mapping.mapping_kind != kind:
            continue
        entity_id = mapping.rack_id if kind == "rack_title" else mapping.device_id
        if entity_id is not None:
            grouped[(mapping.sheet_name, mapping.source_range)].append(entity_id)
    return {
        key: entity_ids[0]
        for key, entity_ids in grouped.items()
        if len(entity_ids) == 1
    }


def _unique_pairs(
    old_keys: dict[str, object],
    new_keys: dict[str, object],
    available_old: set[str],
    available_new: set[str],
) -> dict[str, str]:
    old_groups: dict[object, list[str]] = defaultdict(list)
    new_groups: dict[object, list[str]] = defaultdict(list)
    for entity_id in available_old:
        old_groups[old_keys[entity_id]].append(entity_id)
    for candidate_id in available_new:
        new_groups[new_keys[candidate_id]].append(candidate_id)
    matches: dict[str, str] = {}
    for key, candidate_ids in new_groups.items():
        entity_ids = old_groups.get(key, [])
        if len(candidate_ids) == 1 and len(entity_ids) == 1:
            matches[candidate_ids[0]] = entity_ids[0]
    return matches


def _rack_signature(rack: Rack) -> tuple[object, ...]:
    return (
        rack.height_u,
        rack.direction,
        len(rack.device_columns),
        rack.right_axis_column is not None,
    )


def _candidate_rack_signature(rack: RackCandidate) -> tuple[object, ...]:
    return (
        rack.height_u,
        rack.direction,
        len(rack.device_columns),
        rack.right_axis_column is not None,
    )


def _assign_racks(
    project: RackProject,
    observations: list[_RackObservation],
) -> tuple[dict[str, str], list[IdentityConflict]]:
    old_racks = {item.rack_id: item for item in project.racks}
    new_by_id = {item.candidate.candidate_id: item for item in observations}
    available_old = set(old_racks)
    available_new = set(new_by_id)
    assignments: dict[str, str] = {}

    def accept(matches: dict[str, str]) -> None:
        for candidate_id, rack_id in matches.items():
            if candidate_id in available_new and rack_id in available_old:
                assignments[candidate_id] = rack_id
                available_new.remove(candidate_id)
                available_old.remove(rack_id)

    accept(
        _unique_pairs(
            {rack_id: rack.rack_name.casefold() for rack_id, rack in old_racks.items()},
            {
                candidate_id: item.candidate.rack_name.casefold()
                for candidate_id, item in new_by_id.items()
            },
            available_old,
            available_new,
        )
    )
    by_range = _range_mappings(project, "rack_title")
    range_matches: dict[str, str] = {}
    for candidate_id in available_new:
        item = new_by_id[candidate_id]
        rack_id = by_range.get((item.sheet_name, item.candidate.title_range))
        if rack_id in available_old:
            range_matches[candidate_id] = rack_id
    accept(range_matches)
    accept(
        _unique_pairs(
            {rack_id: _rack_signature(rack) for rack_id, rack in old_racks.items()},
            {
                candidate_id: _candidate_rack_signature(item.candidate)
                for candidate_id, item in new_by_id.items()
            },
            available_old,
            available_new,
        )
    )

    conflicts: list[IdentityConflict] = []
    old_groups: dict[tuple[object, ...], list[str]] = defaultdict(list)
    new_groups: dict[tuple[object, ...], list[str]] = defaultdict(list)
    for rack_id in available_old:
        old_groups[_rack_signature(old_racks[rack_id])].append(rack_id)
    for candidate_id in available_new:
        new_groups[_candidate_rack_signature(new_by_id[candidate_id].candidate)].append(
            candidate_id
        )
    for signature, candidate_ids in new_groups.items():
        entity_ids = old_groups.get(signature, [])
        if entity_ids and (len(entity_ids) > 1 or len(candidate_ids) > 1):
            conflicts.append(
                _conflict(
                    "ambiguous-rack-identity",
                    "Multiple racks share the remaining identity evidence; refusing to guess",
                    entity_ids=sorted(entity_ids),
                    candidate_refs=sorted(candidate_ids),
                    evidence=[f"signature={signature!r}"],
                )
            )
    return assignments, conflicts


def _device_signature(
    device: Device,
    placement: Placement | None,
) -> tuple[object, ...]:
    return (
        device.style_signature,
        placement.height_u if placement is not None else None,
    )


def _candidate_device_signature(item: _DeviceObservation) -> tuple[object, ...]:
    return (item.device.style_signature, item.placement.height_u)


def _assign_devices(
    project: RackProject,
    observations: list[_DeviceObservation],
) -> tuple[dict[str, str], list[IdentityConflict]]:
    old_devices = {item.device_id: item for item in project.devices}
    old_placements = {
        item.device_id: item
        for item in project.placements
        if item.status in {"active", "missing"}
    }
    new_by_id = {item.device.candidate_id: item for item in observations}
    available_old = set(old_devices)
    available_new = set(new_by_id)
    assignments: dict[str, str] = {}

    def accept(matches: dict[str, str]) -> None:
        for candidate_id, device_id in matches.items():
            if candidate_id in available_new and device_id in available_old:
                assignments[candidate_id] = device_id
                available_new.remove(candidate_id)
                available_old.remove(device_id)

    accept(
        _unique_pairs(
            {
                device_id: device.display_text.casefold()
                for device_id, device in old_devices.items()
            },
            {
                candidate_id: item.device.display_text.casefold()
                for candidate_id, item in new_by_id.items()
            },
            available_old,
            available_new,
        )
    )

    by_range = _range_mappings(project, "device")
    conflicts: list[IdentityConflict] = []
    for candidate_id in sorted(available_new):
        item = new_by_id[candidate_id]
        mapped = by_range.get((item.sheet_name, item.device.source_range))
        if mapped is not None and mapped not in available_old:
            conflicts.append(
                _conflict(
                    "competing-device-identity",
                    "Device text and source-range evidence point to different identities",
                    entity_ids=[mapped],
                    candidate_refs=[candidate_id],
                    evidence=[
                        f"source={item.sheet_name}!{item.device.source_range}",
                        f"display_text={item.device.display_text}",
                    ],
                )
            )
    range_matches: dict[str, str] = {}
    for candidate_id in available_new:
        item = new_by_id[candidate_id]
        mapped = by_range.get((item.sheet_name, item.device.source_range))
        if mapped in available_old:
            range_matches[candidate_id] = mapped
    accept(range_matches)

    accept(
        _unique_pairs(
            {
                device_id: _device_signature(
                    device,
                    old_placements.get(device_id),
                )
                for device_id, device in old_devices.items()
            },
            {
                candidate_id: _candidate_device_signature(item)
                for candidate_id, item in new_by_id.items()
            },
            available_old,
            available_new,
        )
    )

    old_groups: dict[tuple[object, ...], list[str]] = defaultdict(list)
    new_groups: dict[tuple[object, ...], list[str]] = defaultdict(list)
    for device_id in available_old:
        old_groups[
            _device_signature(old_devices[device_id], old_placements.get(device_id))
        ].append(device_id)
    for candidate_id in available_new:
        new_groups[_candidate_device_signature(new_by_id[candidate_id])].append(
            candidate_id
        )
    for signature, candidate_ids in new_groups.items():
        entity_ids = old_groups.get(signature, [])
        if entity_ids and (len(entity_ids) > 1 or len(candidate_ids) > 1):
            conflicts.append(
                _conflict(
                    "ambiguous-device-identity",
                    "Multiple devices share the remaining identity evidence; refusing to guess",
                    entity_ids=sorted(entity_ids),
                    candidate_refs=sorted(candidate_ids),
                    evidence=[f"signature={signature!r}"],
                )
            )
    return assignments, _dedupe_conflicts(conflicts)


def _failed_rescan(
    project: RackProject,
    conflicts: list[IdentityConflict],
) -> RescanResult:
    return RescanResult(
        project=project,
        accepted=False,
        conflicts=tuple(_dedupe_conflicts(conflicts)),
    )


def _rescan_from_analysis(
    snapshot_path: Path,
    real_source_path: Path,
    project: RackProject,
    *,
    id_factory: IdFactory,
) -> RescanResult:
    profile = load_stored_profile(project.profile_id, project.metadata)
    analysis = (
        _analyze_workbook_with_profile(snapshot_path, profile)
        if profile is not None
        else analyze_workbook(snapshot_path)
    )
    fingerprint = fingerprint_workbook(snapshot_path, analysis)
    rack_observations, device_observations, observation_conflicts = _observations(analysis)
    blocking = [
        item
        for item in _analysis_conflicts(analysis) + observation_conflicts
        if item.severity == "error"
    ]
    if blocking:
        return _failed_rescan(project, blocking)

    rack_assignments, rack_conflicts = _assign_racks(project, rack_observations)
    device_assignments, device_conflicts = _assign_devices(project, device_observations)
    if rack_conflicts or device_conflicts:
        return _failed_rescan(project, rack_conflicts + device_conflicts)

    old_racks = {item.rack_id: item for item in project.racks}
    old_devices = {item.device_id: item for item in project.devices}
    old_placements = {
        item.device_id: item
        for item in project.placements
        if item.status in {"active", "missing"}
    }
    old_rack_mappings = _entity_mappings(project, "rack_title")
    old_device_mappings = _entity_mappings(project, "device")

    created_racks: list[str] = []
    rack_id_by_candidate: dict[str, str] = {}
    new_racks: list[Rack] = []
    new_mappings: list[SourceMapping] = []
    assigned_rack_ids: set[str] = set()
    for item in rack_observations:
        candidate_id = item.candidate.candidate_id
        rack_id = rack_assignments.get(candidate_id)
        if rack_id is None:
            rack_id = id_factory("RACK")
            created_racks.append(rack_id)
        rack_id_by_candidate[candidate_id] = rack_id
        assigned_rack_ids.add(rack_id)
        new_racks.append(
            _make_rack(item.candidate, rack_id, item.sheet_name, project.profile_id)
        )
        old_mapping = old_rack_mappings.get(rack_id)
        new_mappings.append(
            _make_mapping(
                old_mapping.mapping_id if old_mapping is not None else id_factory("MAP"),
                fingerprint.workbook_sha256,
                item.sheet_name,
                item.candidate.title_range,
                "rack_title",
                rack_id=rack_id,
                confidence=item.candidate.confidence,
            )
        )

    missing_rack_ids = tuple(
        rack_id for rack_id in old_racks if rack_id not in assigned_rack_ids
    )
    for rack_id in missing_rack_ids:
        new_racks.append(replace(old_racks[rack_id], status="missing"))
        mapping = old_rack_mappings.get(rack_id)
        if mapping is not None:
            new_mappings.append(mapping)

    created_devices: list[str] = []
    updated_devices: list[str] = []
    unchanged_devices: list[str] = []
    assigned_device_ids: set[str] = set()
    new_devices: list[Device] = []
    new_placements: list[Placement] = []
    for device_observation in device_observations:
        candidate_id = device_observation.device.candidate_id
        device_id = device_assignments.get(candidate_id)
        old_device = old_devices.get(device_id) if device_id is not None else None
        if device_id is None:
            device_id = id_factory("DEV")
            created_devices.append(device_id)
        assigned_device_ids.add(device_id)
        rack_id = rack_id_by_candidate[device_observation.rack_candidate_id]
        old_placement = old_placements.get(device_id)
        new_device = _make_device(device_observation.device, device_id, old_device)
        new_placement = Placement(
            placement_id=(
                old_placement.placement_id
                if old_placement is not None
                else id_factory("PLC")
            ),
            device_id=device_id,
            rack_id=rack_id,
            start_u=device_observation.placement.start_u,
            end_u=device_observation.placement.end_u,
            status="active",
        )
        new_devices.append(new_device)
        new_placements.append(new_placement)
        old_mapping = old_device_mappings.get(device_id)
        new_mappings.append(
            _make_mapping(
                old_mapping.mapping_id if old_mapping is not None else id_factory("MAP"),
                fingerprint.workbook_sha256,
                device_observation.sheet_name,
                device_observation.device.source_range,
                "device",
                device_id=device_id,
                rack_id=rack_id,
                confidence=device_observation.device.confidence,
                style_signature=device_observation.device.style_signature,
            )
        )
        if old_device is not None:
            changed = (
                old_device.display_text != new_device.display_text
                or old_device.style_signature != new_device.style_signature
                or old_placement is None
                or old_placement.rack_id != new_placement.rack_id
                or old_placement.start_u != new_placement.start_u
                or old_placement.end_u != new_placement.end_u
                or old_placement.status != "active"
            )
            (updated_devices if changed else unchanged_devices).append(device_id)

    missing_device_ids = tuple(
        device_id for device_id in old_devices if device_id not in assigned_device_ids
    )
    for device_id in missing_device_ids:
        new_devices.append(old_devices[device_id])
        old_placement = old_placements.get(device_id)
        if old_placement is not None:
            new_placements.append(replace(old_placement, status="missing"))
        mapping = old_device_mappings.get(device_id)
        if mapping is not None:
            new_mappings.append(mapping)

    rebuilt = RackProject(
        project_id=project.project_id,
        source_workbook=str(normalize_path(real_source_path)),
        workbook_fingerprint=fingerprint.workbook_sha256,
        layout_fingerprint=fingerprint.layout_sha256,
        profile_id=project.profile_id,
        racks=new_racks,
        devices=new_devices,
        placements=new_placements,
        mappings=new_mappings,
        conflicts=[],
        metadata=dict(project.metadata),
    )
    rebuilt = _with_validation(rebuilt, _analysis_conflicts(analysis))
    errors = [item for item in rebuilt.conflicts if item.severity == "error"]
    if errors:
        return _failed_rescan(project, errors)
    return RescanResult(
        project=rebuilt,
        accepted=True,
        conflicts=tuple(rebuilt.conflicts),
        created_rack_ids=tuple(created_racks),
        missing_rack_ids=missing_rack_ids,
        created_device_ids=tuple(created_devices),
        updated_device_ids=tuple(updated_devices),
        unchanged_device_ids=tuple(unchanged_devices),
        missing_device_ids=missing_device_ids,
    )


def _rescan_snapshot(
    snapshot_path: Path,
    real_source_path: Path,
    project: RackProject,
    *,
    id_factory: IdFactory | None = None,
) -> RescanResult:
    return _rescan_from_analysis(
        normalize_path(snapshot_path),
        normalize_path(real_source_path),
        project,
        id_factory=id_factory or default_id_factory,
    )


def rescan_workbook(
    workbook_path: Path,
    project: RackProject,
    *,
    id_factory: IdFactory | None = None,
) -> RescanResult:
    source = normalize_path(workbook_path)
    if not source.is_file():
        raise FileNotFoundError(source)
    existing_errors = project_error_conflicts(project)
    if existing_errors:
        return _failed_rescan(project, existing_errors)
    if project.source_workbook is not None:
        expected = normalize_path(Path(project.source_workbook))
        if source != expected and sha256_file(source) != project.workbook_fingerprint:
            return _failed_rescan(
                project,
                [
                    _conflict(
                        "wrong-source-workbook",
                        "Rescan refused because the workbook is not the bound project source",
                        evidence=[f"expected={expected}", f"actual={source}"],
                    )
                ],
            )
    return _rescan_from_analysis(
        source,
        source,
        project,
        id_factory=id_factory or default_id_factory,
    )

from __future__ import annotations

import os
from copy import copy
from dataclasses import asdict, dataclass, field, replace
from pathlib import Path
from typing import Any, Literal

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.workbook import Workbook as OpenpyxlWorkbook

from racktool.core.backup import create_backup, create_temp_copy
from racktool.core.identity import normalize_path, sha256_file
from racktool.core.project import (
    _rescan_snapshot,
    project_error_conflicts,
)
from racktool.models.domain import Device, Placement, Rack
from racktool.models.project import IdentityConflict, RackProject, SourceMapping

WriteStatus = Literal["applied", "rejected", "failed"]


@dataclass(frozen=True, slots=True)
class WriteAction:
    device_id: str
    rack_id: str
    sheet_name: str
    old_range: str
    new_range: str
    start_u: int
    end_u: int
    display_text: str

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True, slots=True)
class WritePlan:
    actions: tuple[WriteAction, ...]
    conflicts: tuple[IdentityConflict, ...] = ()
    project_id: str | None = None
    source_path: str | None = None
    source_fingerprint: str | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "actions": [item.to_dict() for item in self.actions],
            "conflicts": [item.to_dict() for item in self.conflicts],
            "project_id": self.project_id,
            "source_path": self.source_path,
            "source_fingerprint": self.source_fingerprint,
        }


@dataclass(frozen=True, slots=True)
class WriteResult:
    status: WriteStatus
    plan: WritePlan
    backup_path: str | None = None
    output_path: str | None = None
    project: RackProject | None = None
    message: str = ""
    errors: tuple[str, ...] = field(default_factory=tuple)

    def to_dict(self) -> dict[str, Any]:
        payload = asdict(self)
        payload["plan"] = self.plan.to_dict()
        if self.project is not None:
            payload["project"] = self.project.to_dict()
        return payload


def _conflict(
    code: str,
    message: str,
    *,
    entity_ids: list[str] | None = None,
    evidence: list[str] | None = None,
) -> IdentityConflict:
    return IdentityConflict(
        code=code,
        severity="error",
        message=message,
        entity_ids=entity_ids or [],
        evidence=evidence or [],
    )


def _plan(
    project: RackProject,
    source: Path | None,
    *,
    actions: tuple[WriteAction, ...] = (),
    conflicts: list[IdentityConflict] | None = None,
) -> WritePlan:
    return WritePlan(
        actions=actions,
        conflicts=tuple(conflicts or []),
        project_id=project.project_id,
        source_path=str(source) if source is not None else None,
        source_fingerprint=project.workbook_fingerprint,
    )


def _u_range(start_u: int, end_u: int) -> range:
    return range(min(start_u, end_u), max(start_u, end_u) + 1)


def _active_placements(project: RackProject, device_id: str) -> list[Placement]:
    return [
        item
        for item in project.placements
        if item.device_id == device_id and item.status == "active"
    ]


def _current_device_mappings(project: RackProject, device_id: str) -> list[SourceMapping]:
    return [
        item
        for item in project.mappings
        if item.mapping_kind == "device"
        and item.device_id == device_id
        and item.workbook_fingerprint == project.workbook_fingerprint
    ]


def _occupied_units(
    project: RackProject,
    rack_id: str,
    exclude_device_id: str,
) -> set[int]:
    occupied: set[int] = set()
    for placement in project.placements:
        if placement.status != "active" or placement.rack_id != rack_id:
            continue
        if placement.device_id == exclude_device_id:
            continue
        occupied.update(_u_range(placement.start_u, placement.end_u))
    return occupied


def _bounds(a1: str) -> tuple[int, int, int, int]:
    min_col, min_row, max_col, max_row = range_boundaries(a1)
    if min_col is None or min_row is None or max_col is None or max_row is None:
        raise ValueError(f"Invalid cell range: {a1}")
    return min_col, min_row, max_col, max_row


def _a1(min_col: int, min_row: int, max_col: int, max_row: int) -> str:
    start = f"{get_column_letter(min_col)}{min_row}"
    end = f"{get_column_letter(max_col)}{max_row}"
    return start if start == end else f"{start}:{end}"


def _ranges_overlap(left: tuple[int, int, int, int], right: tuple[int, int, int, int]) -> bool:
    left_min_col, left_min_row, left_max_col, left_max_row = left
    right_min_col, right_min_row, right_max_col, right_max_row = right
    return not (
        left_max_col < right_min_col
        or left_min_col > right_max_col
        or left_max_row < right_min_row
        or left_min_row > right_max_row
    )


def _cell_in_bounds(
    row: int,
    column: int,
    bounds: tuple[int, int, int, int],
) -> bool:
    min_col, min_row, max_col, max_row = bounds
    return min_row <= row <= max_row and min_col <= column <= max_col


def _action_scopes(
    bounds: tuple[int, int, int, int],
    old_bounds: tuple[int, int, int, int],
    new_bounds: tuple[int, int, int, int],
) -> list[str]:
    scopes: list[str] = []
    if _ranges_overlap(bounds, old_bounds):
        scopes.append("source")
    if _ranges_overlap(bounds, new_bounds):
        scopes.append("target")
    return scopes


def _defined_name_bounds(a1: str) -> tuple[int, int, int, int]:
    min_col, min_row, max_col, max_row = range_boundaries(a1)
    return (
        min_col if min_col is not None else 1,
        min_row if min_row is not None else 1,
        max_col if max_col is not None else 16_384,
        max_row if max_row is not None else 1_048_576,
    )


def _defined_name_conflicts(
    workbook: OpenpyxlWorkbook,
    sheet: Any,
    action: WriteAction,
    old_bounds: tuple[int, int, int, int],
    new_bounds: tuple[int, int, int, int],
) -> list[IdentityConflict]:
    conflicts: list[IdentityConflict] = []
    containers = (
        ("workbook", workbook.defined_names),
        ("worksheet", getattr(sheet, "defined_names", {})),
    )
    for name_scope, container in containers:
        for defined_name in container.values():
            try:
                destinations = list(defined_name.destinations)
            except (AttributeError, TypeError, ValueError):
                destinations = []
            attr_text = getattr(defined_name, "attr_text", None)
            if (
                not destinations
                and name_scope == "worksheet"
                and getattr(defined_name, "type", None) == "RANGE"
                and isinstance(attr_text, str)
            ):
                try:
                    _defined_name_bounds(attr_text)
                except ValueError:
                    pass
                else:
                    destinations = [(action.sheet_name, attr_text)]
            for destination_sheet, destination_range in destinations:
                if destination_sheet != action.sheet_name:
                    continue
                try:
                    destination_bounds = _defined_name_bounds(destination_range)
                except (TypeError, ValueError):
                    continue
                scopes = _action_scopes(destination_bounds, old_bounds, new_bounds)
                if not scopes:
                    continue
                conflicts.append(
                    _conflict(
                        "unsupported-defined-name",
                        (
                            "The write action intersects an Excel defined name whose "
                            "meaning Safe Sync cannot update safely"
                        ),
                        entity_ids=[action.device_id, action.rack_id],
                        evidence=[
                            f"defined_name={defined_name.name}",
                            f"name_scope={name_scope}",
                            f"range={destination_range}",
                            f"scope={'+'.join(scopes)}",
                        ],
                    )
                )
    return conflicts


def _unsupported_action_metadata_conflicts(
    workbook: OpenpyxlWorkbook,
    sheet: Any,
    action: WriteAction,
    old_bounds: tuple[int, int, int, int],
    new_bounds: tuple[int, int, int, int],
) -> list[IdentityConflict]:
    conflicts: list[IdentityConflict] = []
    action_cells = {
        (row, column)
        for min_col, min_row, max_col, max_row in (old_bounds, new_bounds)
        for row in range(min_row, max_row + 1)
        for column in range(min_col, max_col + 1)
    }
    for row, column in sorted(action_cells):
        cell = sheet.cell(row, column)
        scopes = []
        if _cell_in_bounds(row, column, old_bounds):
            scopes.append("source")
        if _cell_in_bounds(row, column, new_bounds):
            scopes.append("target")
        coordinate = cell.coordinate
        for metadata_kind, code in (
            ("comment", "unsupported-cell-comment"),
            ("hyperlink", "unsupported-cell-hyperlink"),
        ):
            if getattr(cell, metadata_kind, None) is None:
                continue
            conflicts.append(
                _conflict(
                    code,
                    (
                        f"The {'/'.join(scopes)} cell {coordinate} contains an Excel "
                        f"{metadata_kind} that Safe Sync cannot move or overwrite safely"
                    ),
                    entity_ids=[action.device_id, action.rack_id],
                    evidence=[
                        f"cell={coordinate}",
                        f"scope={'+'.join(scopes)}",
                        f"metadata={metadata_kind}",
                    ],
                )
            )

    data_validations = getattr(sheet, "data_validations", None)
    validations = (
        getattr(data_validations, "dataValidation", ())
        if data_validations is not None
        else ()
    )
    for validation in validations:
        sqref = getattr(validation, "sqref", None)
        if sqref is None:
            continue
        for validation_range in sorted(sqref.ranges, key=str):
            validation_bounds = (
                validation_range.min_col,
                validation_range.min_row,
                validation_range.max_col,
                validation_range.max_row,
            )
            scopes = _action_scopes(validation_bounds, old_bounds, new_bounds)
            if not scopes:
                continue
            conflicts.append(
                _conflict(
                    "unsupported-data-validation",
                    (
                        "The write action intersects Excel data validation that Safe Sync "
                        "cannot relocate or reinterpret safely"
                    ),
                    entity_ids=[action.device_id, action.rack_id],
                    evidence=[
                        f"range={validation_range}",
                        f"scope={'+'.join(scopes)}",
                        "metadata=data-validation",
                    ],
                )
            )
    conflicts.extend(
        _defined_name_conflicts(
            workbook,
            sheet,
            action,
            old_bounds,
            new_bounds,
        )
    )
    return conflicts


def _target_range(
    old_range: str,
    old_rack: Rack,
    target_rack: Rack,
    start_u: int,
    end_u: int,
) -> str:
    old_min_col, _old_min_row, old_max_col, _old_max_row = _bounds(old_range)
    old_columns = list(range(old_min_col, old_max_col + 1))
    source_columns = sorted(old_rack.device_columns)
    target_columns = sorted(target_rack.device_columns)
    width = len(old_columns)
    offsets = [
        offset
        for offset in range(len(source_columns) - width + 1)
        if source_columns[offset : offset + width] == old_columns
    ]
    if len(offsets) != 1:
        raise ValueError("The current device width/lane cannot be mapped unambiguously")
    offset = offsets[0]
    columns = target_columns[offset : offset + width]
    if len(columns) != width or columns != list(range(columns[0], columns[0] + width)):
        raise ValueError("The target rack cannot preserve the device column width/lane")
    rows = [target_rack.u_to_row[unit] for unit in _u_range(start_u, end_u)]
    if sorted(rows) != list(range(min(rows), max(rows) + 1)):
        raise ValueError("The target U rows are not contiguous")
    return _a1(columns[0], min(rows), columns[-1], max(rows))


def _source_conflicts(project: RackProject, source: Path) -> list[IdentityConflict]:
    conflicts: list[IdentityConflict] = []
    if project.source_workbook is None or project.workbook_fingerprint is None:
        conflicts.append(
            _conflict(
                "unbound-project-source",
                "The project has no bound source workbook and fingerprint",
            )
        )
        return conflicts
    expected = normalize_path(Path(project.source_workbook))
    if source != expected:
        conflicts.append(
            _conflict(
                "wrong-source-workbook",
                "The workbook is not the source bound to this project",
                evidence=[f"expected={expected}", f"actual={source}"],
            )
        )
        return conflicts
    if not source.is_file():
        conflicts.append(
            _conflict(
                "source-workbook-missing",
                "The bound source workbook does not exist",
                evidence=[str(source)],
            )
        )
        return conflicts
    current = sha256_file(source)
    if current != project.workbook_fingerprint:
        conflicts.append(
            _conflict(
                "stale-project-fingerprint",
                "The workbook changed after the project state was captured; rescan first",
                evidence=[
                    f"project={project.workbook_fingerprint}",
                    f"current={current}",
                ],
            )
        )
    return conflicts


def _workbook_layout_conflicts(
    source: Path,
    action: WriteAction,
) -> list[IdentityConflict]:
    conflicts: list[IdentityConflict] = []
    workbook = load_workbook(source, read_only=False, data_only=False)
    try:
        if action.sheet_name not in workbook.sheetnames:
            return [
                _conflict(
                    "source-sheet-missing",
                    "The device source Sheet no longer exists",
                    entity_ids=[action.device_id],
                )
            ]
        sheet = workbook[action.sheet_name]
        old_bounds = _bounds(action.old_range)
        new_bounds = _bounds(action.new_range)
        conflicts.extend(
            _unsupported_action_metadata_conflicts(
                workbook,
                sheet,
                action,
                old_bounds,
                new_bounds,
            )
        )
        normalized_old = action.old_range.replace("$", "")
        for merged in sheet.merged_cells.ranges:
            merged_range = str(merged).replace("$", "")
            merged_bounds = (merged.min_col, merged.min_row, merged.max_col, merged.max_row)
            if _ranges_overlap(old_bounds, merged_bounds) and merged_range != normalized_old:
                conflicts.append(
                    _conflict(
                        "source-merge-conflict",
                        "The current device range intersects an unrelated merged range",
                        entity_ids=[action.device_id],
                        evidence=[merged_range],
                    )
                )
            if _ranges_overlap(new_bounds, merged_bounds) and merged_range != normalized_old:
                conflicts.append(
                    _conflict(
                        "target-merge-conflict",
                        "The target range intersects an existing merged range",
                        entity_ids=[action.device_id, action.rack_id],
                        evidence=[merged_range],
                    )
                )
        old_min_col, old_min_row, old_max_col, old_max_row = old_bounds
        new_min_col, new_min_row, new_max_col, new_max_row = new_bounds
        old_cells = {
            (row, column)
            for row in range(old_min_row, old_max_row + 1)
            for column in range(old_min_col, old_max_col + 1)
        }
        for row in range(new_min_row, new_max_row + 1):
            for column in range(new_min_col, new_max_col + 1):
                if (row, column) in old_cells:
                    continue
                value = sheet.cell(row, column).value
                if value not in (None, ""):
                    conflicts.append(
                        _conflict(
                            "target-cell-not-empty",
                            "The target range contains workbook data unknown to the project",
                            entity_ids=[action.device_id, action.rack_id],
                            evidence=[sheet.cell(row, column).coordinate],
                        )
                    )
        anchor_value = sheet.cell(old_min_row, old_min_col).value
        if str(anchor_value) != action.display_text:
            conflicts.append(
                _conflict(
                    "source-device-text-mismatch",
                    "The source cell no longer contains the expected device text",
                    entity_ids=[action.device_id],
                    evidence=[
                        f"expected={action.display_text!r}",
                        f"actual={anchor_value!r}",
                    ],
                )
            )
    finally:
        workbook.close()
    return conflicts


def plan_device_move(
    project: RackProject,
    device_id: str,
    rack_id: str,
    start_u: int,
    end_u: int,
    *,
    workbook_path: Path | None = None,
) -> WritePlan:
    source = (
        normalize_path(workbook_path)
        if workbook_path is not None
        else (
            normalize_path(Path(project.source_workbook))
            if project.source_workbook is not None
            else None
        )
    )
    conflicts = list(project_error_conflicts(project))
    if source is None:
        conflicts.append(
            _conflict("unbound-project-source", "The project has no bound source workbook")
        )
        return _plan(project, source, conflicts=conflicts)
    conflicts.extend(_source_conflicts(project, source))

    devices = [item for item in project.devices if item.device_id == device_id]
    racks = [item for item in project.racks if item.rack_id == rack_id]
    placements = _active_placements(project, device_id)
    mappings = _current_device_mappings(project, device_id)
    if (
        len(devices) != 1
        or len(racks) != 1
        or len(placements) != 1
        or len(mappings) != 1
    ):
        conflicts.append(
            _conflict(
                "ambiguous-project-entity",
                "Device, rack, active Placement, and current Mapping must each be unique",
                entity_ids=[device_id, rack_id],
                evidence=[
                    f"devices={len(devices)}",
                    f"racks={len(racks)}",
                    f"placements={len(placements)}",
                    f"mappings={len(mappings)}",
                ],
            )
        )
        return _plan(project, source, conflicts=conflicts)

    device = devices[0]
    rack = racks[0]
    placement = placements[0]
    mapping = mappings[0]
    old_racks = [item for item in project.racks if item.rack_id == placement.rack_id]
    if len(old_racks) != 1:
        conflicts.append(
            _conflict(
                "ambiguous-source-rack",
                "The current Placement does not resolve to exactly one rack",
                entity_ids=[device_id, placement.rack_id],
            )
        )
        return _plan(project, source, conflicts=conflicts)
    old_rack = old_racks[0]

    if rack.status != "active":
        conflicts.append(
            _conflict(
                "target-rack-missing",
                "A device cannot be moved to a missing rack",
                entity_ids=[device_id, rack_id],
            )
        )
    if mapping.rack_id != placement.rack_id:
        conflicts.append(
            _conflict(
                "mapping-placement-rack-mismatch",
                "The current Mapping and Placement disagree on rack identity",
                entity_ids=[device_id, placement.rack_id, mapping.rack_id or ""],
            )
        )
    if abs(end_u - start_u) + 1 != placement.height_u:
        conflicts.append(
            _conflict(
                "device-height-change",
                "A move must preserve the device U height",
                entity_ids=[device_id],
                evidence=[
                    f"current={placement.height_u}",
                    f"target={abs(end_u - start_u) + 1}",
                ],
            )
        )
    if start_u < 1 or end_u < 1 or start_u > rack.height_u or end_u > rack.height_u:
        conflicts.append(
            _conflict(
                "u-out-of-bounds",
                f"Target U {start_u}-{end_u} is outside 1-{rack.height_u}",
                entity_ids=[device_id, rack_id],
            )
        )
    missing_u = [unit for unit in _u_range(start_u, end_u) if unit not in rack.u_to_row]
    if missing_u:
        conflicts.append(
            _conflict(
                "u-axis-gap",
                f"Rack {rack_id} is missing U rows for {missing_u}",
                entity_ids=[device_id, rack_id],
            )
        )
    overlap = _occupied_units(project, rack_id, device_id).intersection(
        _u_range(start_u, end_u)
    )
    if overlap:
        conflicts.append(
            _conflict(
                "target-u-occupied",
                f"Target U {sorted(overlap)} is already occupied",
                entity_ids=[device_id, rack_id],
            )
        )
    if mapping.sheet_name != rack.source_sheet:
        conflicts.append(
            _conflict(
                "cross-sheet-move-unsupported",
                "V0.4 only writes moves inside the current source Sheet",
                entity_ids=[device_id, rack_id],
            )
        )
    if conflicts:
        return _plan(project, source, conflicts=conflicts)

    try:
        new_range = _target_range(
            mapping.source_range,
            old_rack,
            rack,
            start_u,
            end_u,
        )
    except (KeyError, ValueError) as error:
        conflicts.append(
            _conflict(
                "device-width-or-layout-conflict",
                str(error),
                entity_ids=[device_id, rack_id],
            )
        )
        return _plan(project, source, conflicts=conflicts)

    if (
        placement.rack_id == rack_id
        and min(placement.start_u, placement.end_u) == min(start_u, end_u)
        and max(placement.start_u, placement.end_u) == max(start_u, end_u)
        and mapping.source_range == new_range
    ):
        return _plan(project, source)

    action = WriteAction(
        device_id=device_id,
        rack_id=rack_id,
        sheet_name=mapping.sheet_name,
        old_range=mapping.source_range,
        new_range=new_range,
        start_u=min(start_u, end_u),
        end_u=max(start_u, end_u),
        display_text=device.display_text,
    )
    try:
        conflicts.extend(_workbook_layout_conflicts(source, action))
    except (KeyError, OSError, ValueError) as error:
        conflicts.append(
            _conflict(
                "workbook-preflight-failed",
                f"Workbook preflight failed: {error}",
                entity_ids=[device_id, rack_id],
            )
        )
    if conflicts:
        return _plan(project, source, conflicts=conflicts)
    return _plan(project, source, actions=(action,))


@dataclass(frozen=True, slots=True)
class _CellSnapshot:
    value: Any
    font: object
    fill: object
    alignment: object
    border: object
    number_format: str
    protection: object


def _snapshot_cell(cell: Any) -> _CellSnapshot:
    return _CellSnapshot(
        value=cell.value,
        font=copy(cell.font),
        fill=copy(cell.fill),
        alignment=copy(cell.alignment),
        border=copy(cell.border),
        number_format=cell.number_format,
        protection=copy(cell.protection),
    )


def _restore_cell(cell: Any, snapshot: _CellSnapshot) -> None:
    cell.font = copy(snapshot.font)
    cell.fill = copy(snapshot.fill)
    cell.alignment = copy(snapshot.alignment)
    cell.border = copy(snapshot.border)
    cell.number_format = snapshot.number_format
    cell.protection = copy(snapshot.protection)
    cell.value = snapshot.value


def _apply_action(workbook: OpenpyxlWorkbook, action: WriteAction) -> None:
    sheet = workbook[action.sheet_name]
    old_min_col, old_min_row, old_max_col, old_max_row = _bounds(action.old_range)
    new_min_col, new_min_row, new_max_col, new_max_row = _bounds(action.new_range)
    old_height = old_max_row - old_min_row + 1
    old_width = old_max_col - old_min_col + 1
    if (
        new_max_row - new_min_row + 1 != old_height
        or new_max_col - new_min_col + 1 != old_width
    ):
        raise ValueError("Write action changed the device cell rectangle")

    snapshots = [
        [
            _snapshot_cell(sheet.cell(old_min_row + row, old_min_col + column))
            for column in range(old_width)
        ]
        for row in range(old_height)
    ]
    value = snapshots[0][0].value
    if value in (None, ""):
        value = action.display_text
    normalized_old = action.old_range.replace("$", "")
    for merged in list(sheet.merged_cells.ranges):
        if str(merged).replace("$", "") == normalized_old:
            sheet.unmerge_cells(str(merged))

    new_cells = {
        (row, column)
        for row in range(new_min_row, new_max_row + 1)
        for column in range(new_min_col, new_max_col + 1)
    }
    for row in range(old_min_row, old_max_row + 1):
        for column in range(old_min_col, old_max_col + 1):
            if (row, column) not in new_cells:
                sheet.cell(row, column).value = None

    for row in range(old_height):
        for column in range(old_width):
            target = sheet.cell(new_min_row + row, new_min_col + column)
            _restore_cell(target, snapshots[row][column])
            target.value = None
    sheet.cell(new_min_row, new_min_col).value = value
    if new_min_row != new_max_row or new_min_col != new_max_col:
        sheet.merge_cells(
            start_row=new_min_row,
            start_column=new_min_col,
            end_row=new_max_row,
            end_column=new_max_col,
        )


def _project_after_plan(project: RackProject, plan: WritePlan) -> RackProject:
    placements = list(project.placements)
    mappings = list(project.mappings)
    for action in plan.actions:
        placements = [
            replace(
                item,
                rack_id=action.rack_id,
                start_u=action.start_u,
                end_u=action.end_u,
            )
            if item.device_id == action.device_id and item.status == "active"
            else item
            for item in placements
        ]
        mappings = [
            replace(
                item,
                rack_id=action.rack_id,
                sheet_name=action.sheet_name,
                source_range=action.new_range,
            )
            if item.mapping_kind == "device"
            and item.device_id == action.device_id
            and item.workbook_fingerprint == project.workbook_fingerprint
            else item
            for item in mappings
        ]
    return replace(project, placements=placements, mappings=mappings, conflicts=[])


def _device(project: RackProject, device_id: str) -> Device:
    return next(item for item in project.devices if item.device_id == device_id)


def _placement(project: RackProject, device_id: str) -> Placement:
    return next(
        item
        for item in project.placements
        if item.device_id == device_id and item.status == "active"
    )


def _mapping(project: RackProject, device_id: str) -> SourceMapping:
    return next(
        item
        for item in project.mappings
        if item.mapping_kind == "device"
        and item.device_id == device_id
        and item.workbook_fingerprint == project.workbook_fingerprint
    )


def _validate_written_workbook(
    path: Path,
    real_source: Path,
    project: RackProject,
    plan: WritePlan,
) -> RackProject:
    expected = _project_after_plan(project, plan)
    result = _rescan_snapshot(path, real_source, expected)
    if not result.accepted:
        raise ValueError("; ".join(item.message for item in result.conflicts))
    rescanned = result.project
    errors = [item for item in rescanned.conflicts if item.severity == "error"]
    if errors:
        raise ValueError("; ".join(item.message for item in errors))
    for action in plan.actions:
        placement = _placement(rescanned, action.device_id)
        mapping = _mapping(rescanned, action.device_id)
        if placement.rack_id != action.rack_id:
            raise ValueError(f"Device {action.device_id} did not stay on the target rack")
        if placement.start_u != action.start_u or placement.end_u != action.end_u:
            raise ValueError(f"Device {action.device_id} did not land on the target U range")
        if mapping.source_range != action.new_range:
            raise ValueError(f"Device {action.device_id} mapping was not updated")
        if _device(rescanned, action.device_id).display_text != action.display_text:
            raise ValueError(f"Device {action.device_id} display text was not preserved")
    return rescanned


def _refresh_plan(
    source: Path,
    project: RackProject,
    plan: WritePlan,
) -> WritePlan:
    conflicts: list[IdentityConflict] = []
    if plan.project_id != project.project_id:
        conflicts.append(_conflict("stale-write-plan", "The plan belongs to another project"))
    if plan.source_path != str(source):
        conflicts.append(
            _conflict("stale-write-plan", "The plan belongs to another workbook")
        )
    if plan.source_fingerprint != project.workbook_fingerprint:
        conflicts.append(
            _conflict("stale-write-plan", "The plan fingerprint is no longer current")
        )
    if not plan.actions:
        conflicts.extend(_source_conflicts(project, source))
    refreshed_actions: list[WriteAction] = []
    seen_devices: set[str] = set()
    for action in plan.actions:
        if action.device_id in seen_devices:
            conflicts.append(
                _conflict(
                    "duplicate-plan-device",
                    "A write plan cannot move the same device more than once",
                    entity_ids=[action.device_id],
                )
            )
        seen_devices.add(action.device_id)
        refreshed = plan_device_move(
            project,
            action.device_id,
            action.rack_id,
            action.start_u,
            action.end_u,
            workbook_path=source,
        )
        conflicts.extend(refreshed.conflicts)
        if len(refreshed.actions) != 1 or refreshed.actions[0] != action:
            if not refreshed.conflicts:
                conflicts.append(
                    _conflict(
                        "stale-write-plan",
                        "The planned workbook action changed during freshness validation",
                        entity_ids=[action.device_id, action.rack_id],
                    )
                )
        else:
            refreshed_actions.append(refreshed.actions[0])
    for index, left in enumerate(plan.actions):
        for right in plan.actions[index + 1 :]:
            if left.rack_id == right.rack_id and set(
                _u_range(left.start_u, left.end_u)
            ).intersection(_u_range(right.start_u, right.end_u)):
                conflicts.append(
                    _conflict(
                        "plan-target-overlap",
                        "Two write actions target overlapping rack units",
                        entity_ids=[
                            left.device_id,
                            right.device_id,
                            left.rack_id,
                        ],
                    )
                )
    return _plan(
        project,
        source,
        actions=tuple(refreshed_actions) if not conflicts else (),
        conflicts=conflicts,
    )


def apply_writeback(
    workbook_path: Path,
    project: RackProject,
    plan: WritePlan,
    *,
    replace_source: bool = True,
) -> WriteResult:
    source = normalize_path(workbook_path)
    if plan.conflicts:
        return WriteResult(
            status="rejected",
            plan=plan,
            output_path=str(source),
            project=project,
            message="Write refused because the plan has conflicts",
            errors=tuple(item.message for item in plan.conflicts),
        )
    if not replace_source:
        rejected = replace(
            plan,
            conflicts=(
                _conflict(
                    "detached-write-unsupported",
                    "Safe Sync only commits to the source workbook bound to the project",
                ),
            ),
        )
        return WriteResult(
            status="rejected",
            plan=rejected,
            output_path=str(source),
            project=project,
            message="Detached write-back is not supported",
            errors=tuple(item.message for item in rejected.conflicts),
        )
    refreshed = _refresh_plan(source, project, plan)
    if refreshed.conflicts:
        return WriteResult(
            status="rejected",
            plan=refreshed,
            output_path=str(source),
            project=project,
            message="Write refused because the plan is stale or unsafe",
            errors=tuple(item.message for item in refreshed.conflicts),
        )
    if not plan.actions:
        return WriteResult(
            status="applied",
            plan=plan,
            output_path=str(source),
            project=project,
            message="No workbook changes were required",
        )

    backup_path: Path | None = None
    temp_path: Path | None = None
    try:
        expected_fingerprint = project.workbook_fingerprint
        if expected_fingerprint is None:
            raise ValueError("The project has no workbook fingerprint")
        backup_path = create_backup(source)
        if (
            sha256_file(backup_path) != expected_fingerprint
            or sha256_file(source) != expected_fingerprint
        ):
            raise ValueError("The source workbook changed while its backup was being created")
        temp_path = create_temp_copy(source)
        if sha256_file(temp_path) != expected_fingerprint:
            raise ValueError("The temporary workbook does not match the bound source")

        workbook = load_workbook(temp_path, read_only=False, data_only=False)
        try:
            sheet_names = list(workbook.sheetnames)
            for action in plan.actions:
                _apply_action(workbook, action)
            if list(workbook.sheetnames) != sheet_names:
                raise ValueError("Write-back tried to change the workbook sheet list")
            workbook.save(temp_path)
        finally:
            workbook.close()

        reloaded = load_workbook(temp_path, read_only=False, data_only=False)
        reloaded.close()
        updated_project = _validate_written_workbook(
            temp_path,
            source,
            project,
            plan,
        )
        if sha256_file(source) != expected_fingerprint:
            raise ValueError("The source workbook changed before the atomic commit")
        os.replace(temp_path, source)
        temp_path = None
        return WriteResult(
            status="applied",
            plan=plan,
            backup_path=str(backup_path),
            output_path=str(source),
            project=updated_project,
            message=(
                "Write-back applied after backup, temporary write, reload, "
                "identity validation, and atomic replace"
            ),
        )
    except Exception as error:  # noqa: BLE001
        if temp_path is not None and temp_path.exists():
            temp_path.unlink()
        return WriteResult(
            status="failed",
            plan=plan,
            backup_path=str(backup_path) if backup_path is not None else None,
            output_path=str(source),
            project=project,
            message="Write-back aborted before replacing the source workbook",
            errors=(str(error),),
        )

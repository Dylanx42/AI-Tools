from __future__ import annotations

import hashlib
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple, range_boundaries

from racktool.core.workbook import scan_workbook
from racktool.models.analysis import (
    AnalysisIssue,
    AxisDirection,
    DeviceCandidate,
    PlacementCandidate,
    RackCandidate,
    SheetAnalysis,
    UAxisCandidate,
    WorkbookAnalysis,
)
from racktool.models.workbook import SheetInfo


def _candidate_id(prefix: str, *parts: str) -> str:
    payload = "\x1f".join(parts).encode("utf-8")
    return f"{prefix}-{hashlib.sha256(payload).hexdigest()[:16]}"


@dataclass(frozen=True, slots=True)
class _MergedRegion:
    a1: str
    min_col: int
    min_row: int
    max_col: int
    max_row: int

    def contains(self, row: int, column: int) -> bool:
        return (
            self.min_row <= row <= self.max_row
            and self.min_col <= column <= self.max_col
        )


def _merged_regions(sheet: SheetInfo) -> list[_MergedRegion]:
    regions = []
    for a1 in sheet.merged_ranges:
        min_col, min_row, max_col, max_row = range_boundaries(a1)
        if min_col is None or min_row is None or max_col is None or max_row is None:
            raise ValueError(f"Invalid merged-cell range: {a1}")
        regions.append(_MergedRegion(a1, min_col, min_row, max_col, max_row))
    return regions


def _integer_u(value: Any) -> int | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return None


def _axis_candidate(
    sheet_name: str,
    column_index: int,
    points: list[tuple[int, int]],
    min_length: int,
) -> UAxisCandidate | None:
    if len(points) < min_length:
        return None
    values = [value for _, value in points]
    if min(values) < 1 or max(values) > 100 or len(set(values)) != len(values):
        return None
    step = values[1] - values[0]
    if step not in (-1, 1):
        return None
    direction: AxisDirection = "ascending" if step == 1 else "descending"
    confidence = min(0.99, 0.75 + len(points) / 200)
    return UAxisCandidate(
        sheet_name=sheet_name,
        column_index=column_index,
        column_letter=get_column_letter(column_index),
        start_row=points[0][0],
        end_row=points[-1][0],
        min_u=min(values),
        max_u=max(values),
        direction=direction,
        u_to_row={value: row for row, value in points},
        confidence=confidence,
        evidence=[
            f"{len(points)} consecutive integer U values",
            f"{direction} by worksheet row",
        ],
    )


def detect_u_axes(sheet: SheetInfo, min_length: int = 4) -> list[UAxisCandidate]:
    """Find contiguous vertical integer sequences without assuming a fixed rack height."""
    points_by_column: dict[int, list[tuple[int, int]]] = {}
    for cell in sheet.cells:
        value = _integer_u(cell.value)
        if value is None:
            continue
        row, column = coordinate_to_tuple(cell.coordinate)
        points_by_column.setdefault(column, []).append((row, value))

    candidates: list[UAxisCandidate] = []
    for column, points in points_by_column.items():
        run: list[tuple[int, int]] = []
        step: int | None = None
        for point in sorted(points):
            if not run:
                run = [point]
                step = None
                continue
            row, value = point
            previous_row, previous_value = run[-1]
            next_step = value - previous_value
            if (
                row == previous_row + 1
                and next_step in (-1, 1)
                and (step is None or step == next_step)
            ):
                run.append(point)
                step = next_step
                continue
            candidate = _axis_candidate(sheet.name, column, run, min_length)
            if candidate is not None:
                candidates.append(candidate)
            run = [point]
            step = None
        candidate = _axis_candidate(sheet.name, column, run, min_length)
        if candidate is not None:
            candidates.append(candidate)

    return sorted(candidates, key=lambda item: (item.start_row, item.column_index))


def _same_axis(left: UAxisCandidate, right: UAxisCandidate) -> bool:
    return (
        left.start_row == right.start_row
        and left.end_row == right.end_row
        and left.direction == right.direction
        and left.u_to_row == right.u_to_row
    )


def detect_racks(sheet: SheetInfo, axes: list[UAxisCandidate]) -> list[RackCandidate]:
    """Form rack candidates only when an axis has an adjacent merged title."""
    regions = _merged_regions(sheet)
    cells = {cell.coordinate: cell for cell in sheet.cells}
    axes_by_column: dict[int, list[UAxisCandidate]] = {}
    for axis in axes:
        axes_by_column.setdefault(axis.column_index, []).append(axis)

    racks: list[RackCandidate] = []
    for left_axis in axes:
        title_regions = [
            region
            for region in regions
            if region.min_col == left_axis.column_index
            and region.max_col > region.min_col
            and region.max_row == left_axis.start_row - 1
        ]
        for title_region in title_regions:
            title_coordinate = f"{get_column_letter(title_region.min_col)}{title_region.min_row}"
            title_cell = cells.get(title_coordinate)
            if title_cell is None or not isinstance(title_cell.value, str):
                continue
            rack_name = title_cell.value.strip()
            if not rack_name:
                continue

            right_axis = None
            for column in (title_region.max_col, title_region.max_col + 1):
                if column <= left_axis.column_index:
                    continue
                matching = next(
                    (
                        axis
                        for axis in axes_by_column.get(column, [])
                        if _same_axis(left_axis, axis)
                    ),
                    None,
                )
                if matching is not None:
                    right_axis = matching
                    break

            device_end = (
                right_axis.column_index - 1 if right_axis is not None else title_region.max_col
            )
            device_columns = list(range(left_axis.column_index + 1, device_end + 1))
            if not device_columns:
                continue

            confidence = left_axis.confidence
            evidence = ["merged rack title immediately above U axis"]
            if right_axis is not None:
                confidence = min(confidence, right_axis.confidence)
                evidence.append("matching right or shared U axis")
            else:
                confidence = min(confidence, 0.90)
                evidence.append("single-axis rack at layout edge")

            racks.append(
                RackCandidate(
                    candidate_id=_candidate_id("rack", sheet.name, title_region.a1),
                    sheet_name=sheet.name,
                    rack_name=rack_name,
                    title_range=title_region.a1,
                    left_axis_column=left_axis.column_index,
                    right_axis_column=(
                        right_axis.column_index if right_axis is not None else None
                    ),
                    device_columns=device_columns,
                    start_row=left_axis.start_row,
                    end_row=left_axis.end_row,
                    height_u=left_axis.max_u - left_axis.min_u + 1,
                    direction=left_axis.direction,
                    u_to_row=dict(left_axis.u_to_row),
                    confidence=confidence,
                    evidence=evidence,
                )
            )

    return sorted(racks, key=lambda item: (item.start_row, item.left_axis_column))


def _source_region(
    regions: list[_MergedRegion], rack: RackCandidate, row: int, column: int
) -> _MergedRegion | None:
    return next(
        (
            region
            for region in regions
            if region.contains(row, column)
            and region.min_row >= rack.start_row
            and region.max_row <= rack.end_row
            and region.min_col in rack.device_columns
            and region.max_col in rack.device_columns
        ),
        None,
    )


def detect_devices(
    sheet: SheetInfo, racks: list[RackCandidate]
) -> tuple[list[DeviceCandidate], list[PlacementCandidate]]:
    """Extract text-bearing device candidates from each detected rack's device area."""
    regions = _merged_regions(sheet)
    devices: list[DeviceCandidate] = []
    placements: list[PlacementCandidate] = []
    for rack in racks:
        row_to_u = {row: u for u, row in rack.u_to_row.items()}
        for cell in sheet.cells:
            row, column = coordinate_to_tuple(cell.coordinate)
            display_text = str(cell.value)
            if (
                column not in rack.device_columns
                or row < rack.start_row
                or row > rack.end_row
                or not display_text.strip()
            ):
                continue
            region = _source_region(regions, rack, row, column)
            min_row = region.min_row if region is not None else row
            max_row = region.max_row if region is not None else row
            if min_row not in row_to_u or max_row not in row_to_u:
                continue
            u_values = (row_to_u[min_row], row_to_u[max_row])
            start_u = min(u_values)
            end_u = max(u_values)
            source_range = region.a1 if region is not None else cell.coordinate
            device_evidence = ["non-empty text inside detected rack device area"]
            if region is not None:
                device_evidence.append("source is a merged-cell anchor")
            confidence = rack.confidence
            if not isinstance(cell.value, str):
                confidence = min(confidence, 0.50)
                device_evidence.append("non-text cell converted to display text")
            device_candidate_id = _candidate_id("device", sheet.name, source_range)
            devices.append(
                DeviceCandidate(
                    candidate_id=device_candidate_id,
                    sheet_name=sheet.name,
                    display_text=display_text,
                    value_type=cell.data_type,
                    source_range=source_range,
                    style_signature=cell.style_signature,
                    confidence=confidence,
                    evidence=device_evidence,
                )
            )
            placements.append(
                PlacementCandidate(
                    candidate_id=_candidate_id(
                        "placement", rack.candidate_id, device_candidate_id
                    ),
                    device_candidate_id=device_candidate_id,
                    rack_candidate_id=rack.candidate_id,
                    start_u=start_u,
                    end_u=end_u,
                    height_u=end_u - start_u + 1,
                    confidence=confidence,
                    evidence=["source rows mapped through detected U axis"],
                )
            )

    return devices, placements


def _analysis_issues(
    axes: list[UAxisCandidate],
    racks: list[RackCandidate],
    devices: list[DeviceCandidate],
    placements: list[PlacementCandidate],
) -> list[AnalysisIssue]:
    issues: list[AnalysisIssue] = []
    referenced_axes = {
        (column, rack.start_row, rack.end_row)
        for rack in racks
        for column in (rack.left_axis_column, rack.right_axis_column)
        if column is not None
    }
    for axis in axes:
        axis_key = (axis.column_index, axis.start_row, axis.end_row)
        if axis_key not in referenced_axes:
            source = f"{axis.column_letter}{axis.start_row}:{axis.column_letter}{axis.end_row}"
            issues.append(
                AnalysisIssue(
                    code="unresolved-u-axis",
                    severity="warning",
                    message="U-axis candidate has no adjacent merged rack title",
                    source_ranges=[source],
                )
            )

    racks_by_name: dict[str, list[RackCandidate]] = {}
    for rack in racks:
        racks_by_name.setdefault(rack.rack_name.casefold(), []).append(rack)
    for duplicate_racks in racks_by_name.values():
        if len(duplicate_racks) < 2:
            continue
        issues.append(
            AnalysisIssue(
                code="duplicate-rack-title",
                severity="warning",
                message="Rack title appears in multiple candidate ranges",
                candidate_ids=[rack.candidate_id for rack in duplicate_racks],
                source_ranges=[rack.title_range for rack in duplicate_racks],
            )
        )

    non_text_devices = [device for device in devices if device.value_type != "s"]
    if non_text_devices:
        issues.append(
            AnalysisIssue(
                code="non-text-device-candidate",
                severity="warning",
                message="Non-text cells in a device area require human classification",
                candidate_ids=[device.candidate_id for device in non_text_devices],
                source_ranges=[device.source_range for device in non_text_devices],
            )
        )

    placements_by_rack: dict[str, list[PlacementCandidate]] = {}
    for placement in placements:
        placements_by_rack.setdefault(placement.rack_candidate_id, []).append(placement)
    for rack_placements in placements_by_rack.values():
        by_u: dict[int, list[PlacementCandidate]] = {}
        for placement in rack_placements:
            for u_number in range(placement.start_u, placement.end_u + 1):
                by_u.setdefault(u_number, []).append(placement)
        conflicting_ids = sorted(
            {
                placement.candidate_id
                for placements_at_u in by_u.values()
                if len(placements_at_u) > 1
                for placement in placements_at_u
            }
        )
        if conflicting_ids:
            issues.append(
                AnalysisIssue(
                    code="overlapping-placement-candidates",
                    severity="error",
                    message="Multiple device candidates occupy the same Rack/U position",
                    candidate_ids=conflicting_ids,
                )
            )

    return issues


def analyze_sheet(sheet: SheetInfo) -> SheetAnalysis:
    axes = detect_u_axes(sheet)
    racks = detect_racks(sheet, axes)
    devices, placements = detect_devices(sheet, racks)
    issues = _analysis_issues(axes, racks, devices, placements)
    return SheetAnalysis(
        name=sheet.name,
        u_axes=axes,
        racks=racks,
        devices=devices,
        placements=placements,
        issues=issues,
    )


def analyze_workbook(path: Path) -> WorkbookAnalysis:
    """Analyze XLSX structure into explainable candidates without modifying the workbook."""
    workbook = scan_workbook(path)
    return WorkbookAnalysis(
        format=workbook.format,
        sheets=[analyze_sheet(sheet) for sheet in workbook.sheets],
    )
